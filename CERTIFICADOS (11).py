import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
import re
import pdfplumber
import os
import pandas as pd
import threading
import time
import pickle
from datetime import datetime
from openpyxl.styles import Alignment

try:
    import win32com.client as win32_client
except Exception:
    win32_client = None

try:
    import winreg
except ModuleNotFoundError:
    winreg = None

CONFIG_FILE = "aseguradoras.bin"

try:
    from PyPDF2 import PdfReader, PdfWriter
    PYPDF2_AVAILABLE = True
except Exception:
    PYPDF2_AVAILABLE = False


# -------------------------------------------------------------------
# UTILIDADES REGISTRO WINDOWS HKCU
REG_BASE_KEY = r"Software\CERTIFICADOS_APP"


def save_registry_value(name: str, value: str):
    if winreg is None:
        return
    try:
        key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, REG_BASE_KEY)
        winreg.SetValueEx(key, name, 0, winreg.REG_SZ, str(value))
        winreg.CloseKey(key)
    except Exception:
        pass


def load_registry_value(name: str, default=None):
    if winreg is None:
        return default
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, REG_BASE_KEY, 0, winreg.KEY_READ)
        val, _ = winreg.QueryValueEx(key, name)
        winreg.CloseKey(key)
        return val
    except FileNotFoundError:
        return default
    except Exception:
        return default


def get_outlook_instance():
    if win32_client is None:
        return None
    try:
        return win32_client.GetActiveObject("Outlook.Application")
    except Exception:
        try:
            return win32_client.Dispatch("Outlook.Application")
        except Exception as e:
            print("⚠️ No se pudo conectar con Outlook:", e)
            return None


class RegexExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Extractor Regex - Seguros")

        self.folder_path = tk.StringVar()
        self.selected_producto = tk.StringVar()
        self.selected_aseguradora = tk.StringVar()
        self.resultados = []
        self.start_time = 0
        self.aseguradoras = {}
        self.passwords = {}

        self.init_envios_vars()
        self.load_config()
        self.create_widgets()

    def create_widgets(self):
        # Notebook principal con pestañas
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill="both")

        # Pestaña principal existente
        frame_principal = tk.Frame(self.notebook)
        self.notebook.add(frame_principal, text="Principal")

        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)

        frame_top = tk.Frame(frame_principal)
        frame_top.pack(pady=8, padx=10, fill="x")

        tk.Label(frame_top, text="Producto:").pack(side="left")
        self.producto_cb = ttk.Combobox(
            frame_top,
            textvariable=self.selected_producto,
            values=list(self.aseguradoras.keys()),
            state="readonly",
            width=15,
        )
        self.producto_cb.pack(side="left", padx=5)
        self.producto_cb.bind("<<ComboboxSelected>>", self.on_producto_change)

        tk.Label(frame_top, text="Aseguradora:").pack(side="left")
        self.aseguradora_cb = ttk.Combobox(
            frame_top,
            textvariable=self.selected_aseguradora,
            state="readonly",
            width=30,
        )
        self.aseguradora_cb.pack(side="left", padx=5)
        self.aseguradora_cb.bind("<<ComboboxSelected>>", self.on_aseguradora_change)

        tk.Button(frame_top, text="⚙ Configurar Regex", command=self.open_config_window).pack(side="left", padx=5)
        tk.Button(frame_top, text="Cargar Excel contraseñas", command=self.load_passwords_from_excel).pack(side="left", padx=5)
        self.label_pw_info = tk.Label(frame_top, text="Contraseñas cargadas: 0")
        self.label_pw_info.pack(side="left", padx=8)

        frame_path = tk.Frame(frame_principal)
        frame_path.pack(pady=5, padx=10, fill="x")
        tk.Entry(frame_path, textvariable=self.folder_path, width=60).pack(side="left", padx=5)
        tk.Button(frame_path, text="Seleccionar Carpeta", command=self.select_folder).pack(side="left")

        frame_btns = tk.Frame(frame_principal)
        frame_btns.pack(pady=10)
        self.btn_extract = tk.Button(frame_btns, text="Extraer PDFs", command=self.start_extraction)
        self.btn_extract.pack(side="left", padx=5)
        self.btn_export = tk.Button(frame_btns, text="Exportar Excel", command=self.export_excel, state="disabled")
        self.btn_export.pack(side="left", padx=5)

        self.btn_expand_excel = tk.Button(
            frame_btns,
            text="Procesar Excel Beneficiarios",
            command=self.procesar_excel,
        )
        self.btn_expand_excel.pack(side="left", padx=5)
        tk.Button(frame_btns, text="Separar PDFs", command=self.separar_pdfs).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Renombrar PDFs", command=self.renombrar_pdfs).pack(side="left", padx=5)

        self.progress = ttk.Progressbar(frame_principal, length=450, mode="determinate")
        self.progress.pack(pady=5)

        self.time_label = tk.Label(frame_principal, text="Tiempo de ejecución: 0.00 s")
        self.time_label.pack()

        tk.Label(frame_principal, text="Resultados:").pack()
        self.text_result = tk.Text(frame_principal, height=18, width=100)
        self.text_result.pack(padx=10, pady=5)

        # Agregar pestaña de Envíos
        self.create_envios_tab(self.notebook)

    # === NUEVA PESTAÑA DE ENVÍOS ==========================================================
    def create_envios_tab(self, notebook):
        """Crea la pestaña completa de Envíos dentro del Notebook principal"""
        self.envios_tab = tk.Frame(notebook)
        notebook.add(self.envios_tab, text="Envíos")

        # ---------------- FRAME SUPERIOR: opciones generales ----------------
        frame_top = tk.LabelFrame(self.envios_tab, text="Configuración general", padx=10, pady=8)
        frame_top.pack(fill="x", padx=10, pady=5)

        tk.Label(frame_top, text="Cuenta:").grid(row=0, column=0, sticky="w")
        self.cb_cuenta = ttk.Combobox(frame_top, width=40, state="readonly")
        self.cb_cuenta.grid(row=0, column=1, padx=5, pady=2)

        tk.Label(frame_top, text="Firma:").grid(row=0, column=2, sticky="w")
        self.cb_firma = ttk.Combobox(frame_top, width=35, state="readonly")
        self.cb_firma.grid(row=0, column=3, padx=5, pady=2)

        tk.Label(frame_top, text="Copia (CC):").grid(row=1, column=0, sticky="w")
        self.txt_cc = tk.Text(frame_top, width=60, height=2)
        self.txt_cc.grid(row=1, column=1, columnspan=3, pady=4, sticky="we")

        # ---------------- FRAME CHECKS Y CAMPOS TEXTUALES ----------------
        frame_checks = tk.LabelFrame(self.envios_tab, text="Configuración del mensaje", padx=10, pady=8)
        frame_checks.pack(fill="x", padx=10, pady=5)

        self.var_incluir_subcarpetas = tk.BooleanVar(value=False)
        self.var_asunto_personal = tk.BooleanVar(value=False)
        self.var_cuerpo_personal = tk.BooleanVar(value=False)
        self.var_incluir_nombre = tk.BooleanVar(value=False)

        ttk.Checkbutton(
            frame_checks,
            text="Incluir subcarpetas en búsqueda de adjuntos",
            variable=self.var_incluir_subcarpetas,
        ).grid(row=0, column=0, sticky="w", pady=2, columnspan=2)
        ttk.Checkbutton(
            frame_checks,
            text="Asunto personal por registro",
            variable=self.var_asunto_personal,
            command=self.toggle_asunto_field,
        ).grid(row=1, column=0, sticky="w", pady=2)
        ttk.Checkbutton(
            frame_checks,
            text="Cuerpo personal por registro",
            variable=self.var_cuerpo_personal,
            command=self.toggle_cuerpo_field,
        ).grid(row=2, column=0, sticky="w", pady=2)
        ttk.Checkbutton(
            frame_checks,
            text="Incluir nombre en el saludo",
            variable=self.var_incluir_nombre,
        ).grid(row=3, column=0, sticky="w", pady=2)

        tk.Label(frame_checks, text="Saludo:").grid(row=0, column=2, sticky="nw")
        self.txt_saludo = tk.Text(frame_checks, width=50, height=3)
        self.txt_saludo.grid(row=0, column=3, padx=5, pady=2)

        tk.Label(frame_checks, text="Asunto global:").grid(row=1, column=2, sticky="nw")
        self.txt_asunto = tk.Text(frame_checks, width=50, height=3)
        self.txt_asunto.grid(row=1, column=3, padx=5, pady=2)

        tk.Label(frame_checks, text="Cuerpo global (HTML permitido):").grid(row=2, column=2, sticky="nw")
        self.txt_cuerpo = tk.Text(frame_checks, width=50, height=10)
        self.txt_cuerpo.grid(row=2, column=3, rowspan=2, padx=5, pady=2)

        # ---------------- FRAME DE ARCHIVOS Y PRUEBAS ----------------
        frame_files = tk.LabelFrame(self.envios_tab, text="Estructura y pruebas", padx=10, pady=8)
        frame_files.pack(fill="x", padx=10, pady=5)

        tk.Button(frame_files, text="📁 Seleccionar estructura Excel", command=self.select_estructura_excel).grid(
            row=0, column=0, padx=5
        )
        tk.Button(frame_files, text="📥 Descargar estructura", command=self.descargar_estructura_excel).grid(
            row=0, column=1, padx=5
        )
        tk.Button(frame_files, text="📂 Seleccionar carpeta adjuntos", command=self.select_carpeta_adjuntos).grid(
            row=0, column=2, padx=5
        )

        self.var_modo_prueba = tk.BooleanVar(value=False)
        ttk.Checkbutton(frame_files, text="Modo Prueba", variable=self.var_modo_prueba).grid(row=1, column=0, sticky="w")
        tk.Label(frame_files, text="Correos de prueba:").grid(row=1, column=1, sticky="w")
        self.txt_correos_prueba = tk.Text(frame_files, width=45, height=2)
        self.txt_correos_prueba.grid(row=1, column=2, padx=5)
        tk.Label(frame_files, text="Cantidad de correos:").grid(row=1, column=3, sticky="w")
        self.entry_cant_pruebas = tk.Entry(frame_files, width=8)
        self.entry_cant_pruebas.insert(0, "10")
        self.entry_cant_pruebas.grid(row=1, column=4, padx=5)

        # ---------------- FRAME BOTONES PRINCIPALES ----------------
        frame_btns = tk.Frame(self.envios_tab)
        frame_btns.pack(pady=8)

        self.btn_enviar = tk.Button(
            frame_btns, text="🚀 Enviar correos", width=20, state="normal", command=self.iniciar_envio
        )
        self.btn_enviar.pack(side="left", padx=5)
        self.btn_detener = tk.Button(
            frame_btns, text="⏹️ Detener", width=15, state="disabled", command=self.detener_envio
        )
        self.btn_detener.pack(side="left", padx=5)
        self.btn_reanudar = tk.Button(
            frame_btns, text="🔄 Reanudar", width=15, state="normal", command=self.reanudar_envio
        )
        self.btn_reanudar.pack(side="left", padx=5)
        self.btn_admin = tk.Button(frame_btns, text="🛠 Administrador", width=15, command=self.open_admin_panel)
        self.btn_admin.pack(side="left", padx=5)

        # ---------------- FRAME PROGRESO Y RESULTADOS ----------------
        frame_progress = tk.Frame(self.envios_tab)
        frame_progress.pack(fill="x", padx=10, pady=5)

        self.progress_envios = ttk.Progressbar(frame_progress, length=400, mode="determinate")
        self.progress_envios.pack(side="left", padx=5)
        self.label_progress = tk.Label(frame_progress, text="0% — 0 enviados")
        self.label_progress.pack(side="left", padx=10)
        self.label_time_envios = tk.Label(frame_progress, text="Tiempo: 00:00")
        self.label_time_envios.pack(side="right", padx=10)

        tk.Label(self.envios_tab, text="Resultados:").pack(anchor="w", padx=10)
        self.text_result_envios = tk.Text(self.envios_tab, height=15, width=120)
        self.text_result_envios.pack(padx=10, pady=5)

    # -------------------------------------------------------------------
    # Métodos auxiliares para visibilidad de campos (según checks)
    def toggle_asunto_field(self):
        if self.var_asunto_personal.get():
            self.txt_asunto.config(state="disabled")
        else:
            self.txt_asunto.config(state="normal")

    def toggle_cuerpo_field(self):
        if self.var_cuerpo_personal.get():
            self.txt_cuerpo.config(state="disabled")
        else:
            self.txt_cuerpo.config(state="normal")

    def reanudar_envio(self):
        if not self.verificar_outlook():
            return
        if not self.cargar_cuentas_outlook():
            return
        if not self.cb_cuenta.get().strip():
            messagebox.showwarning("Falta cuenta", "Debes seleccionar una cuenta de Outlook.")
            return
        if not self.cb_firma.get().strip():
            messagebox.showwarning("Falta firma", "Debes seleccionar una firma de Outlook.")
            return
        pendientes = self.cargar_log_parcial()
        if not pendientes:
            return
        if self.envio_en_progreso:
            messagebox.showinfo("Proceso activo", "Ya hay un envío en ejecución.")
            return

        self.envio_en_progreso = True
        self.envio_detener = False
        self.start_time_envio = time.time()
        self.text_result_envios.insert(tk.END, "\n🔄 Reanudando envío desde log parcial...\n")
        self.btn_enviar.config(state="disabled")
        self.btn_reanudar.config(state="disabled")
        self.btn_detener.config(state="normal")

        threading.Thread(target=self.proceso_reanudar, args=(pendientes,), daemon=True).start()
        self.root.after(1000, self.update_tiempo_envios)

    def open_admin_panel(self):
        login = tk.Toplevel(self.root)
        login.title("Acceso Administrador")
        login.geometry("300x180")
        login.resizable(False, False)

        tk.Label(login, text="Usuario:").pack(pady=5)
        user_entry = tk.Entry(login, width=30)
        user_entry.pack()

        tk.Label(login, text="Contraseña:").pack(pady=5)
        pass_entry = tk.Entry(login, width=30, show="*")
        pass_entry.pack()

        def validar_login():
            if user_entry.get().strip().lower() == "admin" and pass_entry.get().strip() == "9530":
                login.destroy()
                self.panel_admin_config()
            else:
                messagebox.showerror("Acceso denegado", "Usuario o contraseña incorrectos.")

        tk.Button(login, text="Entrar", command=validar_login).pack(pady=10)

    # -------------------------------------------------------------------
    # PANEL ADMINISTRADOR (LOGIN Y CONFIGURACIÓN)
    def panel_admin_config(self):
        win = tk.Toplevel(self.root)
        win.title("Configuración Administrador")
        win.geometry("500x300")
        win.resizable(False, False)

        tk.Label(
            win,
            text="Correos de destino (separar con ';'):",
            font=("Segoe UI", 10, "bold"),
        ).pack(pady=5)
        self.txt_admin_correos = tk.Text(win, width=60, height=3)
        self.txt_admin_correos.pack(padx=10)

        correos_guardados = load_registry_value("AdminCorreos", "")
        if correos_guardados:
            self.txt_admin_correos.insert("1.0", correos_guardados)

        auto_log_guardado = load_registry_value("AdminAutoLog", "False")
        auto_activo = str(auto_log_guardado).lower() == "true"

        self.var_admin_auto = tk.BooleanVar(value=auto_activo)
        ttk.Checkbutton(
            win,
            text="Activar envío automático de logs al finalizar procesos",
            variable=self.var_admin_auto,
        ).pack(pady=5)

        def guardar_admin():
            correos = self.txt_admin_correos.get("1.0", "end").strip()
            save_registry_value("AdminCorreos", correos)
            save_registry_value("AdminAutoLog", str(self.var_admin_auto.get()))
            messagebox.showinfo(
                "Guardado", "Configuración del administrador almacenada correctamente."
            )
            win.destroy()

        tk.Button(win, text="Guardar", command=guardar_admin).pack(pady=10)

    # -------------------------------------------------------------------
    # ENVÍO AUTOMÁTICO DE LOG AL ADMIN
    def enviar_log_admin(self, log_path):
        """Envía el log al administrador usando la cuenta seleccionada cuando sea posible."""
        if not log_path:
            return False

        try:
            correos = load_registry_value("AdminCorreos", "")
            auto = load_registry_value("AdminAutoLog", "False")
            if not correos or str(auto).lower() != "true":
                return False

            outlook = get_outlook_instance()
            if outlook is None:
                messagebox.showwarning(
                    "Outlook", "No se pudo conectar con la instancia de Outlook."
                )
                return False

            mail = self._create_mail_item(outlook)
            if mail is None:
                return False
            mail.To = correos
            mail.Subject = (
                f"[Reporte automático] Resultado de envío — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            )
            mail.HTMLBody = (
                f"<p>Se adjunta el resultado del envío ejecutado el "
                f"{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}.</p>"
            )
            mail.Attachments.Add(log_path)

            cuenta_seleccionada = (self.cb_cuenta.get() or "").strip()
            cuenta_asignada = None

            if cuenta_seleccionada:
                try:
                    namespace = outlook.GetNamespace("MAPI")
                    for account in getattr(namespace, "Accounts", []):
                        label = self._format_outlook_account(account)
                        if label == cuenta_seleccionada or cuenta_seleccionada in label:
                            cuenta_asignada = account
                            break
                except Exception as account_err:
                    print(
                        f"⚠️ No se pudo evaluar las cuentas de Outlook para el log: {account_err}"
                    )

            if cuenta_asignada is not None:
                try:
                    mail.SendUsingAccount = cuenta_asignada
                except Exception as assign_err:
                    print(
                        "⚠️ No se pudo asignar la cuenta específica "
                        f"({cuenta_seleccionada}): {assign_err}"
                    )

            try:
                mail.Send()
                self.text_result_envios.insert(
                    tk.END,
                    f"📤 Log enviado al administrador: {correos}"
                    f" {'(cuenta predeterminada)' if cuenta_asignada is None else ''}\n",
                )
                return True
            except Exception as send_err:
                self.text_result_envios.insert(
                    tk.END, f"⚠️ No se pudo enviar log al admin: {send_err}\n"
                )
                return False

        except Exception as e:
            self.text_result_envios.insert(
                tk.END, f"⚠️ No se pudo enviar log al admin: {e}\n"
            )
            return False

    # -------------------------------------------------------------------
    # REANUDAR ENVÍO DESDE LOG PARCIAL
    def cargar_log_parcial(self):
        path = filedialog.askopenfilename(
            title="Seleccionar log parcial", filetypes=[("Archivos Excel", "*.xlsx")]
        )
        if not path:
            return None
        try:
            df_omitidos = pd.read_excel(path, sheet_name="Omitidos")
            pendientes = df_omitidos[["Fila", "Correo"]].values.tolist()
            enviados = len(pd.read_excel(path, sheet_name="Enviados"))
            total = enviados + len(pendientes)
            resumen = (
                f"Enviados: {enviados}\nPendientes: {len(pendientes)}\nTotal: {total}"
            )
            if not pendientes:
                messagebox.showinfo("Sin pendientes", "Todos los correos ya fueron enviados.")
                return None
            if not messagebox.askyesno("Reanudar envío", f"{resumen}\n\n¿Deseas continuar?"):
                return None
            self.modo_reanudar = True
            return pendientes
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el log parcial:\n{e}")
            return None

    # -------------------------------------------------------------------
    def proceso_reanudar(self, pendientes):
        cuenta = self.cb_cuenta.get().strip()
        if not cuenta:
            messagebox.showwarning("Falta cuenta", "Debes seleccionar una cuenta de Outlook.")
            self.reset_envio_estado()
            return

        cc = self.txt_cc.get("1.0", "end").strip()
        saludo = self.txt_saludo.get("1.0", "end").strip()
        asunto = self.txt_asunto.get("1.0", "end").strip()
        cuerpo = self.txt_cuerpo.get("1.0", "end").strip()

        enviados = 0
        omitidos = 0
        errores = 0
        total = len(pendientes)
        self.progress_envios["maximum"] = total

        outlook = get_outlook_instance()
        if outlook is None:
            messagebox.showwarning(
                "Outlook", "No se pudo conectar con la instancia de Outlook abierta."
            )
            self.reset_envio_estado()
            return

        for idx, (fila, correo) in enumerate(pendientes, start=1):
            if self.envio_detener:
                self.text_result_envios.insert(
                    tk.END, "\n⏹️ Reanudación detenida por el usuario.\n"
                )
                break
            if not correo or "@" not in correo:
                omitidos += 1
                self.update_progress(idx, total, enviados, omitidos, errores)
                continue
            try:
                mail = self._create_mail_item(outlook)
                if mail is None:
                    omitidos += 1
                    errores += 1
                    self.text_result_envios.insert(
                        tk.END, f"⚠️ No se pudo preparar el correo para {correo}\n"
                    )
                    self.update_progress(idx, total, enviados, omitidos, errores)
                    continue

                mail.To = correo
                if cc:
                    mail.CC = cc
                mail.Subject = asunto
                cuerpo_html = f"<p>{saludo}</p>{cuerpo}"
                mail.HTMLBody = cuerpo_html + mail.HTMLBody
                mail.Send()
                enviados += 1
                self.text_result_envios.insert(tk.END, f"📧 Reenviado a {correo}\n")
            except Exception as e:
                omitidos += 1
                errores += 1
                self.text_result_envios.insert(
                    tk.END, f"⚠️ Error al reenviar {correo}: {e}\n"
                )
            self.update_progress(idx, total, enviados, omitidos, errores)

        messagebox.showinfo(
            "Reanudación finalizada",
            f"Correos reenviados: {enviados}\nOmitidos: {omitidos}",
        )
        self.reset_envio_estado()

    def on_tab_changed(self, event):
        tab_id = event.widget.index("current")
        if event.widget.tab(tab_id, "text") == "Envíos":
            if self.verificar_outlook():
                self.cargar_cuentas_outlook()

    # -------------------------------------------------------------------
    # VARIABLES INTERNAS DE CONTROL
    def init_envios_vars(self):
        self.envio_detener = False
        self.envio_en_progreso = False
        self.start_time_envio = 0
        self.total_envios = 0
        self.total_omitidos = 0
        self.total_errores = 0
        self.enviados_ok = []
        self.omitidos_log = []
        self.errores_generales = []
        self.folder_adjuntos = ""
        self.estructura_excel = ""
        self.modo_reanudar = False

    # -------------------------------------------------------------------
    # SELECCIÓN DE ARCHIVOS Y CARPETAS
    def select_estructura_excel(self):
        path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel con estructura",
            filetypes=[("Archivos Excel", "*.xlsx *.xls")],
        )
        if path:
            self.estructura_excel = path
            messagebox.showinfo("Estructura seleccionada", f"Archivo cargado:\n{path}")

    def select_carpeta_adjuntos(self):
        folder = filedialog.askdirectory(title="Seleccionar carpeta con adjuntos")
        if folder:
            self.folder_adjuntos = folder
            messagebox.showinfo("Carpeta seleccionada", f"Carpeta base:\n{folder}")

    # -------------------------------------------------------------------
    # DESCARGA DE ESTRUCTURA SEGÚN LOS CHECKS
    def descargar_estructura_excel(self):
        cols = ["Correo", "Identificador"]
        if self.var_asunto_personal.get():
            cols.insert(1, "Asunto")
        if self.var_cuerpo_personal.get():
            cols.insert(1 if "Asunto" not in cols else 2, "Cuerpo")
        if self.var_incluir_nombre.get():
            cols.insert(1, "Nombre")

        df = pd.DataFrame(columns=cols)
        save_path = filedialog.asksaveasfilename(
            title="Guardar estructura como",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Estructura_Envios_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx",
        )
        if save_path:
            df.to_excel(save_path, index=False)
            try:
                os.startfile(save_path)
            except AttributeError:
                messagebox.showinfo("Estructura guardada", f"Archivo generado en:\n{save_path}")

    # -------------------------------------------------------------------
    # VERIFICAR OUTLOOK AL ENTRAR EN PESTAÑA
    def verificar_outlook(self):
        outlook = get_outlook_instance()
        if outlook is None:
            messagebox.showwarning(
                "Outlook", "No se pudo conectar con la instancia de Outlook abierta."
            )
            return False
        try:
            outlook.GetNamespace("MAPI")
        except Exception:
            messagebox.showwarning(
                "Outlook", "No se pudo conectar con la instancia de Outlook abierta."
            )
            return False

        return True

    def cargar_cuentas_outlook(self):
        """Carga las cuentas activas de Outlook, usando fallback por buzones visibles."""
        outlook = get_outlook_instance()
        if outlook is None:
            messagebox.showwarning(
                "Outlook", "No se pudo conectar con la instancia de Outlook abierta."
            )
            return False

        try:
            namespace = outlook.GetNamespace("MAPI")
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo acceder al espacio MAPI:\n{exc}")
            return False

        cuentas = []

        # 🔹 Primer intento: Accounts
        try:
            for account in namespace.Accounts:
                label = self._format_outlook_account(account)
                if label:
                    cuentas.append(label)
        except Exception:
            pass

        # 🔹 Segundo intento: Folders
        if not cuentas:
            try:
                for i in range(1, namespace.Folders.Count + 1):
                    folder = namespace.Folders.Item(i)
                    if folder and folder.Name:
                        cuentas.append(folder.Name)
            except Exception:
                pass

        cuentas = list(dict.fromkeys([c for c in cuentas if c]))
        if not cuentas:
            messagebox.showwarning(
                "Outlook", "No se detectaron cuentas o buzones activos en Outlook."
            )
            return False

        def actualizar_combo():
            self.cb_cuenta["values"] = cuentas
            if self.cb_cuenta.get() not in cuentas:
                self.cb_cuenta.set(cuentas[0])
            self._populate_outlook_signatures()

        self.root.after(0, actualizar_combo)
        return True

    def _unique_ordered(self, items):
        seen = set()
        ordered = []
        for item in items:
            if item and item not in seen:
                seen.add(item)
                ordered.append(item)
        return ordered

    def _populate_outlook_signatures(self):
        firmas = []
        firmas_path = os.path.join(
            os.environ.get("APPDATA", ""), "Microsoft", "Signatures"
        )
        if firmas_path and os.path.isdir(firmas_path):
            try:
                for entry in os.listdir(firmas_path):
                    full_path = os.path.join(firmas_path, entry)
                    if os.path.isdir(full_path):
                        firmas.append(entry)
                    else:
                        base, ext = os.path.splitext(entry)
                        if ext.lower() in {".htm", ".html", ".rtf", ".txt"}:
                            firmas.append(base)
            except Exception:
                firmas = []

        if firmas:
            firmas = self._unique_ordered(firmas)
            self.cb_firma["values"] = firmas
            if self.cb_firma.get() not in firmas:
                self.cb_firma.set(firmas[0])
        else:
            self.cb_firma["values"] = []
            self.cb_firma.set("")

    def _format_outlook_account(self, account):
        """Devuelve una etiqueta segura para mostrar la cuenta, evitando errores COM."""
        if not account:
            return ""

        try:
            display = getattr(account, "DisplayName", "") or getattr(
                account, "UserName", ""
            )
        except Exception:
            display = ""

        try:
            smtp = getattr(account, "SmtpAddress", "")
        except Exception:
            smtp = ""

        display = str(display).strip()
        smtp = str(smtp).strip()

        if display and smtp and display.lower() not in smtp.lower():
            label = f"{display} <{smtp}>"
        else:
            label = smtp or display

        return (label or "").strip()

    def _create_mail_item(self, outlook):
        """Crea un MailItem desde la cuenta seleccionada usando la bandeja de entrada."""
        cuenta_nombre = (self.cb_cuenta.get() or "").strip()

        try:
            namespace = outlook.GetNamespace("MAPI")
        except Exception as exc:
            messagebox.showerror("Outlook", f"No se pudo acceder al espacio MAPI:\n{exc}")
            return None

        mail = None
        cuenta_lower = cuenta_nombre.lower()

        if cuenta_lower:
            try:
                for account in getattr(namespace, "Accounts", []):
                    label = (self._format_outlook_account(account) or "").lower()
                    if cuenta_lower in label:
                        try:
                            inbox = account.DeliveryStore.GetDefaultFolder(6)
                            mail = inbox.Items.Add("IPM.Note")
                        except Exception as folder_err:
                            print(
                                f"⚠️ No se pudo crear correo desde el buzón de {cuenta_nombre}: {folder_err}"
                            )
                            try:
                                mail = outlook.CreateItem(0)
                            except Exception as create_err:
                                print(
                                    f"⚠️ No se pudo crear correo con Outlook.CreateItem: {create_err}"
                                )
                                mail = None

                        if mail is not None:
                            try:
                                mail.SendUsingAccount = account
                            except Exception as assign_err:
                                print(
                                    "⚠️ No se pudo asignar la cuenta específica "
                                    f"({cuenta_nombre}): {assign_err}"
                                )
                        break
            except Exception as accounts_err:
                print(f"⚠️ Error al iterar cuentas de Outlook: {accounts_err}")

        if mail is None:
            if cuenta_lower:
                messagebox.showwarning(
                    "Outlook",
                    f"No se encontró la cuenta '{cuenta_nombre}'.",
                )
                return None

            try:
                inbox = namespace.GetDefaultFolder(6)
                mail = inbox.Items.Add("IPM.Note")
            except Exception:
                try:
                    mail = outlook.CreateItem(0)
                except Exception as create_err:
                    messagebox.showerror(
                        "Outlook",
                        f"No se pudo crear un nuevo correo en Outlook:\n{create_err}",
                    )
                    return None

        return mail

    # -------------------------------------------------------------------
    # INICIAR ENVÍO
    def iniciar_envio(self):
        if not self.verificar_outlook():
            return
        if not self.cargar_cuentas_outlook():
            return
        if self.envio_en_progreso:
            messagebox.showinfo("En proceso", "Ya hay un envío en ejecución.")
            return

        if not self.estructura_excel:
            messagebox.showwarning("Falta estructura", "Debes seleccionar el Excel con los datos de envío.")
            return
        if not self.folder_adjuntos:
            messagebox.showwarning("Falta carpeta", "Debes seleccionar la carpeta con los adjuntos.")
            return
        if not self.cb_cuenta.get().strip():
            messagebox.showwarning("Falta cuenta", "Debes seleccionar una cuenta de Outlook.")
            return
        if not self.cb_firma.get().strip():
            messagebox.showwarning("Falta firma", "Debes seleccionar una firma de Outlook.")
            return

        self.envio_detener = False
        self.envio_en_progreso = True
        self.start_time_envio = time.time()
        self.total_envios = 0
        self.total_omitidos = 0
        self.total_errores = 0
        self.enviados_ok = []
        self.omitidos_log = []
        self.errores_generales = []
        self.btn_enviar.config(state="disabled")
        self.btn_detener.config(state="normal")
        self.btn_reanudar.config(state="disabled")
        self.text_result_envios.delete("1.0", tk.END)

        threading.Thread(target=self.proceso_envio, daemon=True).start()
        self.root.after(1000, self.update_tiempo_envios)

    # -------------------------------------------------------------------
    def detener_envio(self):
        if not self.envio_en_progreso:
            return
        self.envio_detener = True

    # -------------------------------------------------------------------
    def update_tiempo_envios(self):
        if self.envio_en_progreso:
            self.root.after(1000, self.update_tiempo_envios)

    # -------------------------------------------------------------------
    def update_progress(self, current, total, enviados, omitidos, errores):
        porc = (current / total) * 100 if total else 0
        elapsed = time.time() - self.start_time_envio if self.start_time_envio else 0
        avg_time = elapsed / current if current else 0
        remaining = int(max(total - current, 0) * avg_time)
        mins, secs = divmod(remaining, 60)
        self.progress_envios["value"] = current
        self.label_progress.config(
            text=f"{porc:.0f}% — {enviados} enviados, {omitidos} omitidos"
        )
        self.label_time_envios.config(text=f"Restante: {mins:02d}:{secs:02d}")
        self.root.update_idletasks()

    # -------------------------------------------------------------------
    def proceso_envio(self):
        try:
            import pandas as pd  # asegurar import local para hilos
        except Exception as import_error:
            self.errores_generales.append(("Importación", str(import_error)))
            messagebox.showerror(
                "Error",
                f"No se pudieron cargar las dependencias:\n{import_error}",
            )
            self.reset_envio_estado()
            return

        try:
            df = pd.read_excel(self.estructura_excel, dtype=str).fillna("")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el Excel:\n{e}")
            self.reset_envio_estado()
            return

        total = len(df.index)
        if total == 0:
            messagebox.showwarning("Estructura vacía", "El archivo Excel no contiene registros.")
            self.reset_envio_estado()
            return

        self.progress_envios["maximum"] = total
        outlook = get_outlook_instance()
        if outlook is None:
            self.errores_generales.append(("Outlook", "No se pudo inicializar la aplicación"))
            messagebox.showwarning(
                "Outlook", "No se pudo conectar con la instancia de Outlook abierta."
            )
            self.reset_envio_estado()
            return
        cuenta = self.cb_cuenta.get().strip()
        firma_sel = self.cb_firma.get().strip()
        cc = self.txt_cc.get("1.0", "end").strip()

        saludo = self.txt_saludo.get("1.0", "end").strip()
        asunto_global = self.txt_asunto.get("1.0", "end").strip()
        cuerpo_global = self.txt_cuerpo.get("1.0", "end").strip()

        modo_prueba = self.var_modo_prueba.get()
        correos_prueba = self.txt_correos_prueba.get("1.0", "end").strip()
        try:
            cant_pruebas = int(self.entry_cant_pruebas.get() or 0)
        except ValueError:
            cant_pruebas = 0

        if modo_prueba and correos_prueba:
            registros = []
            destinatarios = [c.strip() for c in correos_prueba.split(";") if c.strip()]
            for dest in destinatarios:
                for _ in range(max(cant_pruebas, 1)):
                    registros.append({"Correo": dest})
        else:
            registros = df.to_dict("records")

        enviados = 0
        omitidos = 0
        errores = 0

        for i, row in enumerate(registros, start=1):
            if self.envio_detener:
                self.text_result_envios.insert(tk.END, "\n⏹️ Proceso detenido por el usuario.\n")
                break

            correo = row.get("Correo", "").strip()
            if not correo or "@" not in correo:
                self.omitidos_log.append((i, correo, "Correo vacío o inválido"))
                omitidos += 1
                self.update_progress(i, total, enviados, omitidos, errores)
                continue

            if self.var_asunto_personal.get():
                if "Asunto" not in row:
                    self.omitidos_log.append((i, correo, "Falta campo Asunto"))
                    omitidos += 1
                    self.update_progress(i, total, enviados, omitidos, errores)
                    continue
                asunto = row["Asunto"].strip()
            else:
                asunto = asunto_global

            if self.var_cuerpo_personal.get():
                if "Cuerpo" not in row:
                    self.omitidos_log.append((i, correo, "Falta campo Cuerpo"))
                    omitidos += 1
                    self.update_progress(i, total, enviados, omitidos, errores)
                    continue
                partes = [f"<p>{p.strip()}</p>" for p in row["Cuerpo"].splitlines() if p.strip()]
                cuerpo = "".join(partes)
            else:
                cuerpo = cuerpo_global

            if self.var_incluir_nombre.get():
                if "Nombre" not in row:
                    self.omitidos_log.append((i, correo, "Falta campo Nombre"))
                    omitidos += 1
                    self.update_progress(i, total, enviados, omitidos, errores)
                    continue
                saludo_final = f"{saludo} {row['Nombre'].strip()},"
            else:
                saludo_final = saludo

            palabra = str(row.get("Identificador", "")).strip()
            adjuntos = self.buscar_adjuntos(palabra)
            if not adjuntos:
                self.omitidos_log.append((i, correo, f"Sin adjuntos para {palabra}"))
                omitidos += 1
                self.update_progress(i, total, enviados, omitidos, errores)
                continue

            try:
                mail = self._create_mail_item(outlook)
                if mail is None:
                    self.omitidos_log.append(
                        (i, correo, "No se pudo preparar el correo para la cuenta seleccionada")
                    )
                    omitidos += 1
                    errores += 1
                    self.update_progress(i, total, enviados, omitidos, errores)
                    continue

                mail.To = correo
                if cc:
                    mail.CC = cc
                mail.Subject = f"{asunto} (PRUEBA)" if modo_prueba else asunto
                cuerpo_html = f"<p>{saludo_final}</p>{cuerpo}"
                mail.HTMLBody = cuerpo_html + mail.HTMLBody

                for adj in adjuntos:
                    mail.Attachments.Add(adj)

                mail.Send()
                enviados += 1
                self.enviados_ok.append(
                    (
                        i,
                        correo,
                        asunto,
                        datetime.now().strftime("%H:%M:%S"),
                        cuenta,
                        "PRUEBA" if modo_prueba else "REAL",
                        [os.path.basename(a) for a in adjuntos],
                    )
                )
                self.text_result_envios.insert(tk.END, f"📧 Enviado a {correo}\n")
            except Exception as e_envio:
                self.omitidos_log.append((i, correo, f"Error de envío: {e_envio}"))
                errores += 1

            self.update_progress(i, total, enviados, omitidos, errores)

        self.total_envios = enviados
        self.total_omitidos = omitidos
        self.total_errores = errores
        log_path = self.generar_log_excel(modo_prueba)
        self.mostrar_resumen_final(modo_prueba, log_path)
        self.reset_envio_estado()

    # -------------------------------------------------------------------
    def buscar_adjuntos(self, palabra):
        resultados = []
        if not palabra or not self.folder_adjuntos:
            return resultados
        incluir_sub = self.var_incluir_subcarpetas.get()
        for root_dir, dirs, files in os.walk(self.folder_adjuntos):
            for f in files:
                if palabra.lower() in f.lower():
                    resultados.append(os.path.join(root_dir, f))
            if not incluir_sub:
                break
        return resultados

    # -------------------------------------------------------------------
    def generar_log_excel(self, modo_prueba):
        from openpyxl import Workbook

        tipo = "PRUEBA" if modo_prueba else "REAL"
        fecha = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"Log_Envios_{tipo}_{fecha}.xlsx"
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Enviados"
        if self.enviados_ok:
            cols = ["Fila", "Correo", "Asunto", "Hora de envío", "Cuenta utilizada", "Tipo de envío"]
            max_adj = max(len(r[-1]) for r in self.enviados_ok)
            cols += [f"Adjunto_{n + 1}" for n in range(max_adj)]
            ws1.append(cols)
            for fila in self.enviados_ok:
                base = list(fila[:-1])
                base += fila[-1]
                ws1.append(base)

        ws2 = wb.create_sheet("Omitidos")
        ws2.append(["Fila", "Correo", "Motivo", "Fecha y hora"])
        for fila, correo, motivo in self.omitidos_log:
            ws2.append([fila, correo, motivo, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        ws3 = wb.create_sheet("Errores generales")
        ws3.append(["Error", "Descripción", "Fecha y hora"])
        for err in self.errores_generales:
            ws3.append([err[0], err[1], datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        save_path = os.path.join(os.getcwd(), filename)
        wb.save(save_path)
        try:
            os.startfile(save_path)
        except AttributeError:
            messagebox.showinfo("Log generado", f"Archivo disponible en:\n{save_path}")
        except Exception:
            pass
        self.text_result_envios.insert(tk.END, f"\n📊 Log generado: {filename}\n")
        return save_path

    # -------------------------------------------------------------------
    def mostrar_resumen_final(self, modo_prueba, log_path=None):
        tipo = "PRUEBA" if modo_prueba else "REAL"
        tiempo_total = time.time() - self.start_time_envio if self.start_time_envio else 0
        mins, secs = divmod(int(tiempo_total), 60)
        mensaje = (
            f"Tipo de envío: {tipo}\n"
            f"Correos enviados: {self.total_envios}\n"
            f"Omitidos: {self.total_omitidos}\n"
            f"Errores: {self.total_errores}\n"
            f"Tiempo total: {mins:02d}:{secs:02d}"
        )

        if not log_path:
            fecha = datetime.now().strftime("%Y-%m-%d_%H-%M")
            filename = f"Log_Envios_{tipo}_{fecha}.xlsx"
            log_path = os.path.join(os.getcwd(), filename)

        if log_path and os.path.exists(log_path):
            try:
                from openpyxl import load_workbook

                wb = load_workbook(log_path)
                ws = wb.create_sheet("Resumen final")
                ws.append(["Métrica", "Valor"])
                ws.append(["Tiempo total", f"{mins:02d}:{secs:02d}"])
                promedio = tiempo_total / max(1, self.total_envios) if self.total_envios else 0
                ws.append(["Promedio por correo", f"{promedio:.2f} s"])
                wb.save(log_path)
            except Exception:
                pass

        if log_path:
            self.enviar_log_admin(log_path)

        resumen_win = tk.Toplevel(self.root)
        resumen_win.title("Resumen de envío")
        resumen_win.geometry("350x260")
        resumen_win.resizable(False, False)

        tk.Label(resumen_win, text="📬 Proceso completado", font=("Segoe UI", 10, "bold")).pack(pady=10)
        tk.Message(resumen_win, text=mensaje, width=320).pack(pady=5)

        def abrir_carpeta():
            try:
                os.startfile(os.path.dirname(log_path) if log_path else os.getcwd())
            except Exception:
                pass

        tk.Button(resumen_win, text="Abrir carpeta", command=abrir_carpeta).pack(pady=10)
        tk.Button(resumen_win, text="Cerrar", command=resumen_win.destroy).pack()

    # -------------------------------------------------------------------
    def reset_envio_estado(self):
        self.envio_en_progreso = False
        self.envio_detener = False
        self.btn_enviar.config(state="normal")
        self.btn_detener.config(state="disabled")
        self.btn_reanudar.config(state="disabled")
        self.progress_envios["value"] = 0
        self.label_progress.config(text="0% — 0 enviados")
        self.label_time_envios.config(text="Tiempo: 00:00")
        self.init_envios_vars()
        self.txt_saludo.delete("1.0", "end")
        self.txt_asunto.delete("1.0", "end")
        self.txt_cuerpo.delete("1.0", "end")
        self.txt_cc.delete("1.0", "end")
        self.txt_correos_prueba.delete("1.0", "end")
        self.entry_cant_pruebas.delete(0, "end")
        self.entry_cant_pruebas.insert(0, "10")
        self.cb_cuenta.set("")
        self.cb_firma.set("")
        self.var_asunto_personal.set(False)
        self.var_cuerpo_personal.set(False)
        self.var_incluir_subcarpetas.set(False)
        self.var_incluir_nombre.set(False)
        self.var_modo_prueba.set(False)
        self.toggle_asunto_field()
        self.toggle_cuerpo_field()
        self.label_time_envios.config(text="Tiempo: 00:00")
        self.text_result_envios.insert(tk.END, "\n🧹 Campos limpiados y contador reiniciado.\n")

        # Agregar pestaña de Envíos
        self.create_envios_tab(self.notebook)

    # === NUEVA PESTAÑA DE ENVÍOS ==========================================================
    def create_envios_tab(self, notebook):
        """Crea la pestaña completa de Envíos dentro del Notebook principal"""
        self.envios_tab = tk.Frame(notebook)
        notebook.add(self.envios_tab, text="Envíos")

        # ---------------- FRAME SUPERIOR: opciones generales ----------------
        frame_top = tk.LabelFrame(self.envios_tab, text="Configuración general", padx=10, pady=8)
        frame_top.pack(fill="x", padx=10, pady=5)

        tk.Label(frame_top, text="Cuenta:").grid(row=0, column=0, sticky="w")
        self.cb_cuenta = ttk.Combobox(frame_top, width=40, state="readonly")
        self.cb_cuenta.grid(row=0, column=1, padx=5, pady=2)

        tk.Label(frame_top, text="Firma:").grid(row=0, column=2, sticky="w")
        self.cb_firma = ttk.Combobox(frame_top, width=35, state="readonly")
        self.cb_firma.grid(row=0, column=3, padx=5, pady=2)

        tk.Label(frame_top, text="Copia (CC):").grid(row=1, column=0, sticky="w")
        self.txt_cc = tk.Text(frame_top, width=60, height=2)
        self.txt_cc.grid(row=1, column=1, columnspan=3, pady=4, sticky="we")

        # ---------------- FRAME CHECKS Y CAMPOS TEXTUALES ----------------
        frame_checks = tk.LabelFrame(self.envios_tab, text="Configuración del mensaje", padx=10, pady=8)
        frame_checks.pack(fill="x", padx=10, pady=5)

        self.var_incluir_subcarpetas = tk.BooleanVar(value=False)
        self.var_asunto_personal = tk.BooleanVar(value=False)
        self.var_cuerpo_personal = tk.BooleanVar(value=False)
        self.var_incluir_nombre = tk.BooleanVar(value=False)

        ttk.Checkbutton(
            frame_checks,
            text="Incluir subcarpetas en búsqueda de adjuntos",
            variable=self.var_incluir_subcarpetas,
        ).grid(row=0, column=0, sticky="w", pady=2, columnspan=2)
        ttk.Checkbutton(
            frame_checks,
            text="Asunto personal por registro",
            variable=self.var_asunto_personal,
            command=self.toggle_asunto_field,
        ).grid(row=1, column=0, sticky="w", pady=2)
        ttk.Checkbutton(
            frame_checks,
            text="Cuerpo personal por registro",
            variable=self.var_cuerpo_personal,
            command=self.toggle_cuerpo_field,
        ).grid(row=2, column=0, sticky="w", pady=2)
        ttk.Checkbutton(
            frame_checks,
            text="Incluir nombre en el saludo",
            variable=self.var_incluir_nombre,
        ).grid(row=3, column=0, sticky="w", pady=2)

        tk.Label(frame_checks, text="Saludo:").grid(row=0, column=2, sticky="nw")
        self.txt_saludo = tk.Text(frame_checks, width=50, height=3)
        self.txt_saludo.grid(row=0, column=3, padx=5, pady=2)

        tk.Label(frame_checks, text="Asunto global:").grid(row=1, column=2, sticky="nw")
        self.txt_asunto = tk.Text(frame_checks, width=50, height=3)
        self.txt_asunto.grid(row=1, column=3, padx=5, pady=2)

        tk.Label(frame_checks, text="Cuerpo global (HTML permitido):").grid(row=2, column=2, sticky="nw")
        self.txt_cuerpo = tk.Text(frame_checks, width=50, height=10)
        self.txt_cuerpo.grid(row=2, column=3, rowspan=2, padx=5, pady=2)

        # ---------------- FRAME DE ARCHIVOS Y PRUEBAS ----------------
        frame_files = tk.LabelFrame(self.envios_tab, text="Estructura y pruebas", padx=10, pady=8)
        frame_files.pack(fill="x", padx=10, pady=5)

        tk.Button(frame_files, text="📁 Seleccionar estructura Excel", command=self.select_estructura_excel).grid(
            row=0, column=0, padx=5
        )
        tk.Button(frame_files, text="📥 Descargar estructura", command=self.descargar_estructura_excel).grid(
            row=0, column=1, padx=5
        )
        tk.Button(frame_files, text="📂 Seleccionar carpeta adjuntos", command=self.select_carpeta_adjuntos).grid(
            row=0, column=2, padx=5
        )

        self.var_modo_prueba = tk.BooleanVar(value=False)
        ttk.Checkbutton(frame_files, text="Modo Prueba", variable=self.var_modo_prueba).grid(row=1, column=0, sticky="w")
        tk.Label(frame_files, text="Correos de prueba:").grid(row=1, column=1, sticky="w")
        self.txt_correos_prueba = tk.Text(frame_files, width=45, height=2)
        self.txt_correos_prueba.grid(row=1, column=2, padx=5)
        tk.Label(frame_files, text="Cantidad de correos:").grid(row=1, column=3, sticky="w")
        self.entry_cant_pruebas = tk.Entry(frame_files, width=8)
        self.entry_cant_pruebas.insert(0, "10")
        self.entry_cant_pruebas.grid(row=1, column=4, padx=5)

        # ---------------- FRAME BOTONES PRINCIPALES ----------------
        frame_btns = tk.Frame(self.envios_tab)
        frame_btns.pack(pady=8)

        self.btn_enviar = tk.Button(
            frame_btns, text="🚀 Enviar correos", width=20, state="normal", command=self.iniciar_envio
        )
        self.btn_enviar.pack(side="left", padx=5)
        self.btn_detener = tk.Button(
            frame_btns, text="⏹️ Detener", width=15, state="disabled", command=self.detener_envio
        )
        self.btn_detener.pack(side="left", padx=5)
        self.btn_reanudar = tk.Button(
            frame_btns, text="🔄 Reanudar", width=15, state="normal", command=self.reanudar_envio
        )
        self.btn_reanudar.pack(side="left", padx=5)
        self.btn_admin = tk.Button(frame_btns, text="🛠 Administrador", width=15, command=self.open_admin_panel)
        self.btn_admin.pack(side="left", padx=5)

        # ---------------- FRAME PROGRESO Y RESULTADOS ----------------
        frame_progress = tk.Frame(self.envios_tab)
        frame_progress.pack(fill="x", padx=10, pady=5)

        self.progress_envios = ttk.Progressbar(frame_progress, length=400, mode="determinate")
        self.progress_envios.pack(side="left", padx=5)
        self.label_progress = tk.Label(frame_progress, text="0% — 0 enviados")
        self.label_progress.pack(side="left", padx=10)
        self.label_time_envios = tk.Label(frame_progress, text="Tiempo: 00:00")
        self.label_time_envios.pack(side="right", padx=10)

        tk.Label(self.envios_tab, text="Resultados:").pack(anchor="w", padx=10)
        self.text_result_envios = tk.Text(self.envios_tab, height=15, width=120)
        self.text_result_envios.pack(padx=10, pady=5)

    # -------------------------------------------------------------------
    # Métodos auxiliares para visibilidad de campos (según checks)
    def toggle_asunto_field(self):
        if self.var_asunto_personal.get():
            self.txt_asunto.config(state="disabled")
        else:
            self.txt_asunto.config(state="normal")

    def toggle_cuerpo_field(self):
        if self.var_cuerpo_personal.get():
            self.txt_cuerpo.config(state="disabled")
        else:
            self.txt_cuerpo.config(state="normal")

    def reanudar_envio(self):
        if not self.verificar_outlook():
            return
        if not self.cargar_cuentas_outlook():
            return
        if not self.cb_cuenta.get().strip():
            messagebox.showwarning("Falta cuenta", "Debes seleccionar una cuenta de Outlook.")
            return
        if not self.cb_firma.get().strip():
            messagebox.showwarning("Falta firma", "Debes seleccionar una firma de Outlook.")
            return
        pendientes = self.cargar_log_parcial()
        if not pendientes:
            return
        if self.envio_en_progreso:
            messagebox.showinfo("Proceso activo", "Ya hay un envío en ejecución.")
            return

        self.envio_en_progreso = True
        self.envio_detener = False
        self.start_time_envio = time.time()
        self.text_result_envios.insert(tk.END, "\n🔄 Reanudando envío desde log parcial...\n")
        self.btn_enviar.config(state="disabled")
        self.btn_reanudar.config(state="disabled")
        self.btn_detener.config(state="normal")

        threading.Thread(target=self.proceso_reanudar, args=(pendientes,), daemon=True).start()
        self.root.after(1000, self.update_tiempo_envios)

    def open_admin_panel(self):
        login = tk.Toplevel(self.root)
        login.title("Acceso Administrador")
        login.geometry("300x180")
        login.resizable(False, False)

        tk.Label(login, text="Usuario:").pack(pady=5)
        user_entry = tk.Entry(login, width=30)
        user_entry.pack()

        tk.Label(login, text="Contraseña:").pack(pady=5)
        pass_entry = tk.Entry(login, width=30, show="*")
        pass_entry.pack()

        def validar_login():
            if user_entry.get().strip().lower() == "admin" and pass_entry.get().strip() == "9530":
                login.destroy()
                self.panel_admin_config()
            else:
                messagebox.showerror("Acceso denegado", "Usuario o contraseña incorrectos.")

        tk.Button(login, text="Entrar", command=validar_login).pack(pady=10)

    # -------------------------------------------------------------------
    # PANEL ADMINISTRADOR (LOGIN Y CONFIGURACIÓN)
    def panel_admin_config(self):
        win = tk.Toplevel(self.root)
        win.title("Configuración Administrador")
        win.geometry("500x300")
        win.resizable(False, False)

        tk.Label(
            win,
            text="Correos de destino (separar con ';'):",
            font=("Segoe UI", 10, "bold"),
        ).pack(pady=5)
        self.txt_admin_correos = tk.Text(win, width=60, height=3)
        self.txt_admin_correos.pack(padx=10)

        correos_guardados = load_registry_value("AdminCorreos", "")
        if correos_guardados:
            self.txt_admin_correos.insert("1.0", correos_guardados)

        auto_log_guardado = load_registry_value("AdminAutoLog", "False")
        auto_activo = str(auto_log_guardado).lower() == "true"

        self.var_admin_auto = tk.BooleanVar(value=auto_activo)
        ttk.Checkbutton(
            win,
            text="Activar envío automático de logs al finalizar procesos",
            variable=self.var_admin_auto,
        ).pack(pady=5)

        def guardar_admin():
            correos = self.txt_admin_correos.get("1.0", "end").strip()
            save_registry_value("AdminCorreos", correos)
            save_registry_value("AdminAutoLog", str(self.var_admin_auto.get()))
            messagebox.showinfo(
                "Guardado", "Configuración del administrador almacenada correctamente."
            )
            win.destroy()

        tk.Button(win, text="Guardar", command=guardar_admin).pack(pady=10)

    # -------------------------------------------------------------------
    # ENVÍO AUTOMÁTICO DE LOG AL ADMIN
    def enviar_log_admin(self, log_path):
        if not log_path:
            return
        try:
            correos = load_registry_value("AdminCorreos", "")
            auto = load_registry_value("AdminAutoLog", "False")
            if not correos or str(auto).lower() != "true":
                return

            outlook = get_outlook_instance()
            if outlook is None:
                raise RuntimeError("No se pudo conectar con la instancia de Outlook abierta.")
            mail = outlook.CreateItem(0)
            mail.To = correos
            mail.Subject = (
                f"[Reporte automático] Resultado de envío — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            )
            mail.HTMLBody = (
                f"<p>Se adjunta el resultado del envío ejecutado el "
                f"{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}.</p>"
            )
            mail.Attachments.Add(log_path)
            mail.Send()
            self.text_result_envios.insert(
                tk.END, f"📤 Log enviado al administrador: {correos}\n"
            )
        except Exception as e:
            self.text_result_envios.insert(
                tk.END, f"⚠️ No se pudo enviar log al admin: {e}\n"
            )

    # -------------------------------------------------------------------
    # REANUDAR ENVÍO DESDE LOG PARCIAL
    def cargar_log_parcial(self):
        path = filedialog.askopenfilename(
            title="Seleccionar log parcial", filetypes=[("Archivos Excel", "*.xlsx")]
        )
        if not path:
            return None
        try:
            df_omitidos = pd.read_excel(path, sheet_name="Omitidos")
            pendientes = df_omitidos[["Fila", "Correo"]].values.tolist()
            enviados = len(pd.read_excel(path, sheet_name="Enviados"))
            total = enviados + len(pendientes)
            resumen = (
                f"Enviados: {enviados}\nPendientes: {len(pendientes)}\nTotal: {total}"
            )
            if not pendientes:
                messagebox.showinfo("Sin pendientes", "Todos los correos ya fueron enviados.")
                return None
            if not messagebox.askyesno("Reanudar envío", f"{resumen}\n\n¿Deseas continuar?"):
                return None
            self.modo_reanudar = True
            return pendientes
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el log parcial:\n{e}")
            return None

    # -------------------------------------------------------------------
    def proceso_reanudar(self, pendientes):
        cuenta = self.cb_cuenta.get().strip()
        if not cuenta:
            messagebox.showwarning("Falta cuenta", "Debes seleccionar una cuenta de Outlook.")
            self.reset_envio_estado()
            return

        cc = self.txt_cc.get("1.0", "end").strip()
        saludo = self.txt_saludo.get("1.0", "end").strip()
        asunto = self.txt_asunto.get("1.0", "end").strip()
        cuerpo = self.txt_cuerpo.get("1.0", "end").strip()

        enviados = 0
        omitidos = 0
        errores = 0
        total = len(pendientes)
        self.progress_envios["maximum"] = total

        outlook = get_outlook_instance()
        if outlook is None:
            messagebox.showwarning(
                "Outlook", "No se pudo conectar con la instancia de Outlook abierta."
            )
            self.reset_envio_estado()
            return

        for idx, (fila, correo) in enumerate(pendientes, start=1):
            if self.envio_detener:
                self.text_result_envios.insert(
                    tk.END, "\n⏹️ Reanudación detenida por el usuario.\n"
                )
                break
            if not correo or "@" not in correo:
                omitidos += 1
                self.update_progress(idx, total, enviados, omitidos, errores)
                continue
            try:
                mail = outlook.CreateItem(0)
                mail.To = correo
                if cc:
                    mail.CC = cc
                mail.Subject = asunto
                cuerpo_html = f"<p>{saludo}</p>{cuerpo}"
                mail.HTMLBody = cuerpo_html + mail.HTMLBody
                mail.Send()
                enviados += 1
                self.text_result_envios.insert(tk.END, f"📧 Reenviado a {correo}\n")
            except Exception as e:
                omitidos += 1
                errores += 1
                self.text_result_envios.insert(
                    tk.END, f"⚠️ Error al reenviar {correo}: {e}\n"
                )
            self.update_progress(idx, total, enviados, omitidos, errores)

        messagebox.showinfo(
            "Reanudación finalizada",
            f"Correos reenviados: {enviados}\nOmitidos: {omitidos}",
        )
        self.reset_envio_estado()

    def on_tab_changed(self, event):
        tab_id = event.widget.index("current")
        if event.widget.tab(tab_id, "text") == "Envíos":
            if self.verificar_outlook():
                self.cargar_cuentas_outlook()

    # -------------------------------------------------------------------
    # VARIABLES INTERNAS DE CONTROL
    def init_envios_vars(self):
        self.envio_detener = False
        self.envio_en_progreso = False
        self.start_time_envio = 0
        self.total_envios = 0
        self.total_omitidos = 0
        self.total_errores = 0
        self.enviados_ok = []
        self.omitidos_log = []
        self.errores_generales = []
        self.folder_adjuntos = ""
        self.estructura_excel = ""
        self.modo_reanudar = False

    # -------------------------------------------------------------------
    # SELECCIÓN DE ARCHIVOS Y CARPETAS
    def select_estructura_excel(self):
        path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel con estructura",
            filetypes=[("Archivos Excel", "*.xlsx *.xls")],
        )
        if path:
            self.estructura_excel = path
            messagebox.showinfo("Estructura seleccionada", f"Archivo cargado:\n{path}")

    def select_carpeta_adjuntos(self):
        folder = filedialog.askdirectory(title="Seleccionar carpeta con adjuntos")
        if folder:
            self.folder_adjuntos = folder
            messagebox.showinfo("Carpeta seleccionada", f"Carpeta base:\n{folder}")

    # -------------------------------------------------------------------
    # DESCARGA DE ESTRUCTURA SEGÚN LOS CHECKS
    def descargar_estructura_excel(self):
        cols = ["Correo", "Identificador"]
        if self.var_asunto_personal.get():
            cols.insert(1, "Asunto")
        if self.var_cuerpo_personal.get():
            cols.insert(1 if "Asunto" not in cols else 2, "Cuerpo")
        if self.var_incluir_nombre.get():
            cols.insert(1, "Nombre")

        df = pd.DataFrame(columns=cols)
        save_path = filedialog.asksaveasfilename(
            title="Guardar estructura como",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Estructura_Envios_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx",
        )
        if save_path:
            df.to_excel(save_path, index=False)
            try:
                os.startfile(save_path)
            except AttributeError:
                messagebox.showinfo("Estructura guardada", f"Archivo generado en:\n{save_path}")

    # -------------------------------------------------------------------
    # VERIFICAR OUTLOOK AL ENTRAR EN PESTAÑA
    def verificar_outlook(self):
        outlook = get_outlook_instance()
        if outlook is None:
            messagebox.showwarning(
                "Outlook", "No se pudo conectar con la instancia de Outlook abierta."
            )
            return False
        try:
            outlook.GetNamespace("MAPI")
        except Exception:
            messagebox.showwarning(
                "Outlook", "No se pudo conectar con la instancia de Outlook abierta."
            )
            return False

        return True

    def cargar_cuentas_outlook(self):
        """Carga las cuentas activas de Outlook, usando fallback por buzones visibles."""
        outlook = get_outlook_instance()
        if outlook is None:
            messagebox.showwarning(
                "Outlook", "No se pudo conectar con la instancia de Outlook abierta."
            )
            return False

        try:
            namespace = outlook.GetNamespace("MAPI")
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo acceder al espacio MAPI:\n{exc}")
            return False

        cuentas = []

        # 🔹 Primer intento: Accounts
        try:
            for account in namespace.Accounts:
                label = self._format_outlook_account(account)
                if label:
                    cuentas.append(label)
        except Exception:
            pass

        # 🔹 Segundo intento: Folders
        if not cuentas:
            try:
                for i in range(1, namespace.Folders.Count + 1):
                    folder = namespace.Folders.Item(i)
                    if folder and folder.Name:
                        cuentas.append(folder.Name)
            except Exception:
                pass

        cuentas = list(dict.fromkeys([c for c in cuentas if c]))
        if not cuentas:
            messagebox.showwarning(
                "Outlook", "No se detectaron cuentas o buzones activos en Outlook."
            )
            return False

        def actualizar_combo():
            self.cb_cuenta["values"] = cuentas
            if self.cb_cuenta.get() not in cuentas:
                self.cb_cuenta.set(cuentas[0])
            self._populate_outlook_signatures()

        self.root.after(0, actualizar_combo)
        return True

    def _unique_ordered(self, items):
        seen = set()
        ordered = []
        for item in items:
            if item and item not in seen:
                seen.add(item)
                ordered.append(item)
        return ordered

    def _populate_outlook_signatures(self):
        firmas = []
        firmas_path = os.path.join(
            os.environ.get("APPDATA", ""), "Microsoft", "Signatures"
        )
        if firmas_path and os.path.isdir(firmas_path):
            try:
                for entry in os.listdir(firmas_path):
                    full_path = os.path.join(firmas_path, entry)
                    if os.path.isdir(full_path):
                        firmas.append(entry)
                    else:
                        base, ext = os.path.splitext(entry)
                        if ext.lower() in {".htm", ".html", ".rtf", ".txt"}:
                            firmas.append(base)
            except Exception:
                firmas = []

        if firmas:
            firmas = self._unique_ordered(firmas)
            self.cb_firma["values"] = firmas
            if self.cb_firma.get() not in firmas:
                self.cb_firma.set(firmas[0])
        else:
            self.cb_firma["values"] = []
            self.cb_firma.set("")

    def _format_outlook_account(self, account):
        """Devuelve una etiqueta segura para mostrar la cuenta, evitando errores COM."""
        if not account:
            return ""

        try:
            display = getattr(account, "DisplayName", "") or getattr(
                account, "UserName", ""
            )
        except Exception:
            display = ""

        try:
            smtp = getattr(account, "SmtpAddress", "")
        except Exception:
            smtp = ""

        display = str(display).strip()
        smtp = str(smtp).strip()

        if display and smtp and display.lower() not in smtp.lower():
            label = f"{display} <{smtp}>"
        else:
            label = smtp or display

        return (label or "").strip()

    # -------------------------------------------------------------------
    # INICIAR ENVÍO
    def iniciar_envio(self):
        if not self.verificar_outlook():
            return
        if not self.cargar_cuentas_outlook():
            return
        if self.envio_en_progreso:
            messagebox.showinfo("En proceso", "Ya hay un envío en ejecución.")
            return

        if not self.estructura_excel:
            messagebox.showwarning("Falta estructura", "Debes seleccionar el Excel con los datos de envío.")
            return
        if not self.folder_adjuntos:
            messagebox.showwarning("Falta carpeta", "Debes seleccionar la carpeta con los adjuntos.")
            return
        if not self.cb_cuenta.get().strip():
            messagebox.showwarning("Falta cuenta", "Debes seleccionar una cuenta de Outlook.")
            return
        if not self.cb_firma.get().strip():
            messagebox.showwarning("Falta firma", "Debes seleccionar una firma de Outlook.")
            return

        self.envio_detener = False
        self.envio_en_progreso = True
        self.start_time_envio = time.time()
        self.total_envios = 0
        self.total_omitidos = 0
        self.total_errores = 0
        self.enviados_ok = []
        self.omitidos_log = []
        self.errores_generales = []
        self.btn_enviar.config(state="disabled")
        self.btn_detener.config(state="normal")
        self.btn_reanudar.config(state="disabled")
        self.text_result_envios.delete("1.0", tk.END)

        threading.Thread(target=self.proceso_envio, daemon=True).start()
        self.root.after(1000, self.update_tiempo_envios)

    # -------------------------------------------------------------------
    def detener_envio(self):
        if not self.envio_en_progreso:
            return
        self.envio_detener = True

    # -------------------------------------------------------------------
    def update_tiempo_envios(self):
        if self.envio_en_progreso:
            self.root.after(1000, self.update_tiempo_envios)

    # -------------------------------------------------------------------
    def update_progress(self, current, total, enviados, omitidos, errores):
        porc = (current / total) * 100 if total else 0
        elapsed = time.time() - self.start_time_envio if self.start_time_envio else 0
        avg_time = elapsed / current if current else 0
        remaining = int(max(total - current, 0) * avg_time)
        mins, secs = divmod(remaining, 60)
        self.progress_envios["value"] = current
        self.label_progress.config(
            text=f"{porc:.0f}% — {enviados} enviados, {omitidos} omitidos"
        )
        self.label_time_envios.config(text=f"Restante: {mins:02d}:{secs:02d}")
        self.root.update_idletasks()

    # -------------------------------------------------------------------
    def proceso_envio(self):
        try:
            import pandas as pd  # asegurar import local para hilos
        except Exception as import_error:
            self.errores_generales.append(("Importación", str(import_error)))
            messagebox.showerror(
                "Error",
                f"No se pudieron cargar las dependencias:\n{import_error}",
            )
            self.reset_envio_estado()
            return

        try:
            df = pd.read_excel(self.estructura_excel, dtype=str).fillna("")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el Excel:\n{e}")
            self.reset_envio_estado()
            return

        total = len(df.index)
        if total == 0:
            messagebox.showwarning("Estructura vacía", "El archivo Excel no contiene registros.")
            self.reset_envio_estado()
            return

        self.progress_envios["maximum"] = total
        outlook = get_outlook_instance()
        if outlook is None:
            self.errores_generales.append(("Outlook", "No se pudo inicializar la aplicación"))
            messagebox.showwarning(
                "Outlook", "No se pudo conectar con la instancia de Outlook abierta."
            )
            self.reset_envio_estado()
            return
        cuenta = self.cb_cuenta.get().strip()
        firma_sel = self.cb_firma.get().strip()
        cc = self.txt_cc.get("1.0", "end").strip()

        saludo = self.txt_saludo.get("1.0", "end").strip()
        asunto_global = self.txt_asunto.get("1.0", "end").strip()
        cuerpo_global = self.txt_cuerpo.get("1.0", "end").strip()

        modo_prueba = self.var_modo_prueba.get()
        correos_prueba = self.txt_correos_prueba.get("1.0", "end").strip()
        try:
            cant_pruebas = int(self.entry_cant_pruebas.get() or 0)
        except ValueError:
            cant_pruebas = 0

        if modo_prueba and correos_prueba:
            registros = []
            destinatarios = [c.strip() for c in correos_prueba.split(";") if c.strip()]
            for dest in destinatarios:
                for _ in range(max(cant_pruebas, 1)):
                    registros.append({"Correo": dest})
        else:
            registros = df.to_dict("records")

        enviados = 0
        omitidos = 0
        errores = 0

        for i, row in enumerate(registros, start=1):
            if self.envio_detener:
                self.text_result_envios.insert(tk.END, "\n⏹️ Proceso detenido por el usuario.\n")
                break

            correo = row.get("Correo", "").strip()
            if not correo or "@" not in correo:
                self.omitidos_log.append((i, correo, "Correo vacío o inválido"))
                omitidos += 1
                self.update_progress(i, total, enviados, omitidos, errores)
                continue

            if self.var_asunto_personal.get():
                if "Asunto" not in row:
                    self.omitidos_log.append((i, correo, "Falta campo Asunto"))
                    omitidos += 1
                    self.update_progress(i, total, enviados, omitidos, errores)
                    continue
                asunto = row["Asunto"].strip()
            else:
                asunto = asunto_global

            if self.var_cuerpo_personal.get():
                if "Cuerpo" not in row:
                    self.omitidos_log.append((i, correo, "Falta campo Cuerpo"))
                    omitidos += 1
                    self.update_progress(i, total, enviados, omitidos, errores)
                    continue
                partes = [f"<p>{p.strip()}</p>" for p in row["Cuerpo"].splitlines() if p.strip()]
                cuerpo = "".join(partes)
            else:
                cuerpo = cuerpo_global

            if self.var_incluir_nombre.get():
                if "Nombre" not in row:
                    self.omitidos_log.append((i, correo, "Falta campo Nombre"))
                    omitidos += 1
                    self.update_progress(i, total, enviados, omitidos, errores)
                    continue
                saludo_final = f"{saludo} {row['Nombre'].strip()},"
            else:
                saludo_final = saludo

            palabra = str(row.get("Identificador", "")).strip()
            adjuntos = self.buscar_adjuntos(palabra)
            if not adjuntos:
                self.omitidos_log.append((i, correo, f"Sin adjuntos para {palabra}"))
                omitidos += 1
                self.update_progress(i, total, enviados, omitidos, errores)
                continue

            try:
                mail = outlook.CreateItem(0)
                mail.To = correo
                if cc:
                    mail.CC = cc
                mail.Subject = f"{asunto} (PRUEBA)" if modo_prueba else asunto
                cuerpo_html = f"<p>{saludo_final}</p>{cuerpo}"
                mail.HTMLBody = cuerpo_html + mail.HTMLBody

                for adj in adjuntos:
                    mail.Attachments.Add(adj)

                mail.Send()
                enviados += 1
                self.enviados_ok.append(
                    (
                        i,
                        correo,
                        asunto,
                        datetime.now().strftime("%H:%M:%S"),
                        cuenta,
                        "PRUEBA" if modo_prueba else "REAL",
                        [os.path.basename(a) for a in adjuntos],
                    )
                )
                self.text_result_envios.insert(tk.END, f"📧 Enviado a {correo}\n")
            except Exception as e_envio:
                self.omitidos_log.append((i, correo, f"Error de envío: {e_envio}"))
                errores += 1

            self.update_progress(i, total, enviados, omitidos, errores)

        self.total_envios = enviados
        self.total_omitidos = omitidos
        self.total_errores = errores
        log_path = self.generar_log_excel(modo_prueba)
        self.mostrar_resumen_final(modo_prueba, log_path)
        self.reset_envio_estado()

    # -------------------------------------------------------------------
    def buscar_adjuntos(self, palabra):
        resultados = []
        if not palabra or not self.folder_adjuntos:
            return resultados
        incluir_sub = self.var_incluir_subcarpetas.get()
        for root_dir, dirs, files in os.walk(self.folder_adjuntos):
            for f in files:
                if palabra.lower() in f.lower():
                    resultados.append(os.path.join(root_dir, f))
            if not incluir_sub:
                break
        return resultados

    # -------------------------------------------------------------------
    def generar_log_excel(self, modo_prueba):
        from openpyxl import Workbook

        tipo = "PRUEBA" if modo_prueba else "REAL"
        fecha = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"Log_Envios_{tipo}_{fecha}.xlsx"
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Enviados"
        if self.enviados_ok:
            cols = ["Fila", "Correo", "Asunto", "Hora de envío", "Cuenta utilizada", "Tipo de envío"]
            max_adj = max(len(r[-1]) for r in self.enviados_ok)
            cols += [f"Adjunto_{n + 1}" for n in range(max_adj)]
            ws1.append(cols)
            for fila in self.enviados_ok:
                base = list(fila[:-1])
                base += fila[-1]
                ws1.append(base)

        ws2 = wb.create_sheet("Omitidos")
        ws2.append(["Fila", "Correo", "Motivo", "Fecha y hora"])
        for fila, correo, motivo in self.omitidos_log:
            ws2.append([fila, correo, motivo, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        ws3 = wb.create_sheet("Errores generales")
        ws3.append(["Error", "Descripción", "Fecha y hora"])
        for err in self.errores_generales:
            ws3.append([err[0], err[1], datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        save_path = os.path.join(os.getcwd(), filename)
        wb.save(save_path)
        try:
            os.startfile(save_path)
        except AttributeError:
            messagebox.showinfo("Log generado", f"Archivo disponible en:\n{save_path}")
        except Exception:
            pass
        self.text_result_envios.insert(tk.END, f"\n📊 Log generado: {filename}\n")
        return save_path

    # -------------------------------------------------------------------
    def mostrar_resumen_final(self, modo_prueba, log_path=None):
        tipo = "PRUEBA" if modo_prueba else "REAL"
        tiempo_total = time.time() - self.start_time_envio if self.start_time_envio else 0
        mins, secs = divmod(int(tiempo_total), 60)
        mensaje = (
            f"Tipo de envío: {tipo}\n"
            f"Correos enviados: {self.total_envios}\n"
            f"Omitidos: {self.total_omitidos}\n"
            f"Errores: {self.total_errores}\n"
            f"Tiempo total: {mins:02d}:{secs:02d}"
        )

        if not log_path:
            fecha = datetime.now().strftime("%Y-%m-%d_%H-%M")
            filename = f"Log_Envios_{tipo}_{fecha}.xlsx"
            log_path = os.path.join(os.getcwd(), filename)

        if log_path and os.path.exists(log_path):
            try:
                from openpyxl import load_workbook

                wb = load_workbook(log_path)
                ws = wb.create_sheet("Resumen final")
                ws.append(["Métrica", "Valor"])
                ws.append(["Tiempo total", f"{mins:02d}:{secs:02d}"])
                promedio = tiempo_total / max(1, self.total_envios) if self.total_envios else 0
                ws.append(["Promedio por correo", f"{promedio:.2f} s"])
                wb.save(log_path)
            except Exception:
                pass

        if log_path:
            self.enviar_log_admin(log_path)

        resumen_win = tk.Toplevel(self.root)
        resumen_win.title("Resumen de envío")
        resumen_win.geometry("350x260")
        resumen_win.resizable(False, False)

        tk.Label(resumen_win, text="📬 Proceso completado", font=("Segoe UI", 10, "bold")).pack(pady=10)
        tk.Message(resumen_win, text=mensaje, width=320).pack(pady=5)

        def abrir_carpeta():
            try:
                os.startfile(os.path.dirname(log_path) if log_path else os.getcwd())
            except Exception:
                pass

        tk.Button(resumen_win, text="Abrir carpeta", command=abrir_carpeta).pack(pady=10)
        tk.Button(resumen_win, text="Cerrar", command=resumen_win.destroy).pack()

    # -------------------------------------------------------------------
    def reset_envio_estado(self):
        self.envio_en_progreso = False
        self.envio_detener = False
        self.btn_enviar.config(state="normal")
        self.btn_detener.config(state="disabled")
        self.btn_reanudar.config(state="disabled")
        self.progress_envios["value"] = 0
        self.label_progress.config(text="0% — 0 enviados")
        self.label_time_envios.config(text="Tiempo: 00:00")
        self.init_envios_vars()
        self.txt_saludo.delete("1.0", "end")
        self.txt_asunto.delete("1.0", "end")
        self.txt_cuerpo.delete("1.0", "end")
        self.txt_cc.delete("1.0", "end")
        self.txt_correos_prueba.delete("1.0", "end")
        self.entry_cant_pruebas.delete(0, "end")
        self.entry_cant_pruebas.insert(0, "10")
        self.cb_cuenta.set("")
        self.cb_firma.set("")
        self.var_asunto_personal.set(False)
        self.var_cuerpo_personal.set(False)
        self.var_incluir_subcarpetas.set(False)
        self.var_incluir_nombre.set(False)
        self.var_modo_prueba.set(False)
        self.toggle_asunto_field()
        self.toggle_cuerpo_field()
        self.label_time_envios.config(text="Tiempo: 00:00")
        self.text_result_envios.insert(tk.END, "\n🧹 Campos limpiados y contador reiniciado.\n")

    def procesar_excel(self):
        path = filedialog.askopenfilename(
            title="Seleccionar Excel a procesar",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not path:
            return
        try:
            df = pd.read_excel(path, dtype=str)

    # 🔹 Normalizamos los nombres de columnas
            df.columns = df.columns.str.strip().str.upper()
            

            if "BENEFICIARIOS" not in df.columns:
                messagebox.showwarning("Advertencia", "El archivo no tiene la columna 'BENEFICIARIOS'.")
                return

            filas_expandidas = []

            for _, row in df.iterrows():
                beneficiarios = str(row.get("BENEFICIARIOS", "")).split("\n")
                for b in beneficiarios:
                    b = b.strip()
                    if not b:
                        continue

                    match = re.match(r"(.+?)\s+([A-ZÁÉÍÓÚÑ]+)\s+(\d+ ?%)", b)
                    if match:
                        nombre = match.group(1).strip()
                        parentesco = match.group(2).strip()
                        porcentaje = match.group(3).strip()

                        nueva_fila = row.to_dict()
                        nueva_fila["BENEFICIARIOS"] = nombre
                        nueva_fila["PARENTESCO"] = parentesco
                        nueva_fila["PORCENTAJE"] = porcentaje
                        filas_expandidas.append(nueva_fila)
                    else:
                        nueva_fila = row.to_dict()
                        nueva_fila["PARENTESCO"] = ""
                        nueva_fila["PORCENTAJE"] = ""
                        filas_expandidas.append(nueva_fila)

            if not filas_expandidas:
                messagebox.showwarning("Advertencia", "No se encontraron beneficiarios para expandir.")
                return

            df_expandido = pd.DataFrame(filas_expandidas)

            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Guardar Excel expandido"
            )
            if save_path:
                df_expandido.to_excel(save_path, index=False)
                messagebox.showinfo("Éxito", f"Excel expandido guardado en:\n{save_path}")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo procesar el archivo:\n{e}")

    def on_producto_change(self, event=None):
        producto = self.selected_producto.get()
        if producto in self.aseguradoras:
            self.aseguradora_cb["values"] = list(self.aseguradoras[producto].keys())
            if self.aseguradoras[producto]:
                self.aseguradora_cb.set(list(self.aseguradoras[producto].keys())[0])
            else:
                self.aseguradora_cb.set("")

    def on_aseguradora_change(self, event=None):
        self.selected_aseguradora.set(self.aseguradora_cb.get())

    def refresh_productos_cb(self):
        self.producto_cb["values"] = list(self.aseguradoras.keys())

    def refresh_aseguradoras_cb(self):
        producto = self.selected_producto.get()
        if producto in self.aseguradoras:
            self.aseguradora_cb["values"] = list(self.aseguradoras[producto].keys())
        else:
            self.aseguradora_cb["values"] = []
        self.aseguradora_cb.set("")

    def select_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.folder_path.set(folder)

    def start_extraction(self):
        if not self.folder_path.get():
            messagebox.showwarning("Advertencia", "Seleccione una carpeta primero.")
            return
        if not self.selected_producto.get() or not self.selected_aseguradora.get():
            messagebox.showwarning("Advertencia", "Seleccione producto y aseguradora.")
            return

        self.btn_extract.config(state="disabled")
        self.btn_export.config(state="disabled")
        self.progress["value"] = 0
        self.text_result.delete("1.0", tk.END)
        self.start_time = time.time()

        t = threading.Thread(target=self.extract_pdfs, daemon=True)
        t.start()
        self.update_time()

    def update_time(self):
        if self.btn_extract["state"] == "disabled":
            elapsed = time.time() - self.start_time
            self.time_label.config(text=f"Tiempo de ejecución: {elapsed:.2f} s")
            self.root.after(100, self.update_time)

    def extract_pdfs(self):
        folder = self.folder_path.get()
        producto = self.selected_producto.get()
        aseguradora = self.selected_aseguradora.get()
        patrones = self.aseguradoras.get(producto, {}).get(aseguradora, {})

        pdf_files = [f for f in os.listdir(folder) if f.lower().endswith(".pdf")]
        total = len(pdf_files)
        self.resultados = []

        if total == 0:
            messagebox.showwarning("Advertencia", "No se encontraron PDFs en la carpeta.")
            self.btn_extract.config(state="normal")
            return

        for i, file in enumerate(pdf_files, 1):
            filepath = os.path.join(folder, file)
            data = {"Archivo": file, "Producto": producto, "Aseguradora": aseguradora}
            try:
                raw_text = self.read_pdf_text(filepath, file)

                # Guardamos una versión normalizada solo para campos problemáticos
                text_clean = raw_text.replace('\n', ' ').replace('\r', ' ')
                text_clean = re.sub(r'\s{2,}', ' ', text_clean).strip()

                pw = self.get_password_for_file(file)
                data["Contraseña"] = pw if pw else ""

                rows = [data]

                for campo, patron in patrones.items():
                    try:
                        normalized_field = campo.strip().lower()
                        campo_mayus = campo.upper()

                        # 🔹 1. Campos de Beneficiarios (una fila por cada coincidencia)
                        if "beneficiario" in normalized_field:
                            text_for_field = raw_text
                            text_for_field = text_for_field.replace("\r\n", "\n").replace("\r", "\n")
                            text_for_field = re.sub(r"(?<=\b[A-ZÁÉÍÓÚÑÜ])\s(?=[A-ZÁÉÍÓÚÑÜ]\b)", "", text_for_field)
                            text_for_field = re.sub(r"(\d)\s+([A-ZÁÉÍÓÚÑÜ])", r"\1\n\2", text_for_field)
                            text_for_field = re.sub(r"\s{3,}", "\n", text_for_field)
                            text_for_field = re.sub(r"\n0+\s*", "\n", text_for_field)
                            text_for_field = re.sub(r"\n{2,}", "\n", text_for_field).strip()
                            text_for_field = re.sub(r"\s{4,}", "\n", text_for_field)
                            text_for_field = re.sub(r"[ \t]{3,}", " ", text_for_field)

                            # 🔹 Limpieza avanzada para Midline
                            text_for_field = re.sub(
                                r"(?<=\b[A-ZÁÉÍÓÚÑÜ])\s(?=[A-ZÁÉÍÓÚÑÜ]\b)",
                                "",
                                text_for_field,
                            )  # Une letras separadas
                            text_for_field = re.sub(
                                r"(\d{1,3})\s+(?=[A-ZÁÉÍÓÚÑÜ])",
                                r"\1\n",
                                text_for_field,
                            )  # Salto tras el porcentaje
                            text_for_field = text_for_field.replace(
                                "Fecha de impresión", "\nFecha de impresión"
                            )  # Delimita el final

                            matches = re.findall(patron, text_for_field, re.MULTILINE | re.IGNORECASE)

                            new_rows = []
                            if matches:
                                for row in rows:
                                    added = False
                                    for match in matches:
                                        if isinstance(match, tuple):
                                            entry = " ".join(part for part in match if part).strip()
                                        else:
                                            entry = str(match).strip()

                                        if entry:
                                            new_row = row.copy()
                                            new_row[campo] = entry
                                            new_rows.append(new_row)
                                            added = True
                                    if not added:
                                        row_copy = row.copy()
                                        row_copy[campo] = ""
                                        new_rows.append(row_copy)
                            else:
                                for row in rows:
                                    row_copy = row.copy()
                                    row_copy[campo] = ""
                                    new_rows.append(row_copy)

                            rows = new_rows if new_rows else rows
                            continue

                        # 🔹 2. Campos de Mascotas (una fila por cada mascota, sin distinción de mayúsculas)
                        if "mascota" in normalized_field:
                            text_for_field = raw_text
                            text_for_field = text_for_field.replace("\r\n", "\n").replace("\r", "\n")
                            text_for_field = re.sub(r"(?<=\b[A-ZÁÉÍÓÚÑÜ])\s(?=[A-ZÁÉÍÓÚÑÜ]\b)", "", text_for_field)
                            text_for_field = re.sub(r"(\d)\s+([A-ZÁÉÍÓÚÑÜ])", r"\1\n\2", text_for_field)
                            text_for_field = re.sub(r"\s{3,}", "\n", text_for_field)
                            text_for_field = re.sub(r"\n0+\s*", "\n", text_for_field)
                            text_for_field = re.sub(r"(\n0+\s*){2,}", "\n", text_for_field)
                            text_for_field = re.sub(r"\n{2,}", "\n", text_for_field)
                            text_for_field = re.sub(r"[ \t]{3,}", " ", text_for_field)

                            matches = re.findall(patron, text_for_field, re.MULTILINE | re.IGNORECASE)

                            new_rows = []
                            if matches:
                                for row in rows:
                                    added = False
                                    for match in matches:
                                        if isinstance(match, tuple):
                                            entry = " ".join(part for part in match if part).strip()
                                        else:
                                            entry = str(match).strip()

                                        if entry:
                                            new_row = row.copy()
                                            new_row[campo] = entry
                                            new_rows.append(new_row)
                                            added = True
                                    if not added:
                                        row_copy = row.copy()
                                        row_copy[campo] = ""
                                        new_rows.append(row_copy)
                            else:
                                for row in rows:
                                    row_copy = row.copy()
                                    row_copy[campo] = ""
                                    new_rows.append(row_copy)

                            rows = new_rows if new_rows else rows
                            continue

                        if "vigencia" in normalized_field or "beneficiario" in normalized_field:
                            source_text = text_clean
                        else:
                            source_text = raw_text

                        if (
                            ("NUMERO" in campo_mayus and "DOCUMENTO" in campo_mayus)
                            or "VALOR ASEGURADO" in campo_mayus
                            or "BENEFICIARIOS" in campo_mayus
                        ):
                            source_text = source_text.replace("  ", "\n")

                        match = re.search(patron, source_text, re.MULTILINE | re.IGNORECASE)
                        if match:
                            if match.lastindex:
                                groups = [match.group(i) for i in range(1, match.lastindex + 1)]
                                first_value = next((g for g in groups if g), "")
                                value = first_value.strip()
                            else:
                                value = match.group(0).strip()
                        else:
                            value = ""

                        if "documento" in normalized_field and value:
                            value = re.sub(r"\D", "", value)

                        for row in rows:
                            row[campo] = value
                    except re.error as re_err:
                        for row in rows:
                            row[campo] = f"Error regex: {re_err}"
                    except Exception as e:
                        for row in rows:
                            row[campo] = f"Error: {e}"

                self.resultados.extend(rows)

            except Exception as e:
                err_text = str(e)
                data["Error"] = err_text
                data["Contraseña"] = f"Error: {err_text}"
                self.resultados.append(data)

            self.progress["value"] = (i / total) * 100
            self.text_result.insert(tk.END, f"Procesado: {file}\n")
            self.root.update_idletasks()

        dur = time.time() - self.start_time
        self.time_label.config(text=f"Tiempo de ejecución: {dur:.2f} s")

        self.text_result.insert(tk.END, "\n=== Resultados completos ===\n")
        for r in self.resultados:
            self.text_result.insert(tk.END, str(r) + "\n")

        self.btn_extract.config(state="normal")
        self.btn_export.config(state="normal")

    def get_password_for_file(self, filename):
        return self.passwords.get(filename)

    def read_pdf_text(self, filepath, filename_for_lookup):
        try:
            with pdfplumber.open(filepath) as pdf:
                return "\n".join([page.extract_text() or "" for page in pdf.pages])
        except Exception as e_direct:
            pw = self.get_password_for_file(filename_for_lookup)
            if not PYPDF2_AVAILABLE:
                raise Exception(f"Error al abrir PDF: {e_direct}")

            try:
                reader = PdfReader(filepath)
                if getattr(reader, "is_encrypted", False):
                    if not pw:
                        raise Exception("PDF encriptado y sin contraseña registrada.")
                    reader.decrypt(pw)

                text_pages = []
                for p in reader.pages:
                    try:
                        t = p.extract_text() or ""
                    except Exception:
                        t = ""
                    text_pages.append(t)
                return "\n".join(text_pages)

            except Exception as e2:
                raise Exception(f"Error al desencriptar/leer {filename_for_lookup}: {e2}")

    def load_passwords_from_excel(self):
        path = filedialog.askopenfilename(
            title="Seleccionar Excel de contraseñas",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            if path.lower().endswith(".csv"):
                df = pd.read_csv(path, dtype=str)
            else:
                df = pd.read_excel(path, dtype=str)

            self.passwords.clear()
            for _, row in df.iterrows():
                self.passwords[str(row.iloc[0]).strip()] = str(row.iloc[1]).strip()

            self.label_pw_info.config(text=f"Contraseñas cargadas: {len(self.passwords)}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo:\n{e}")

    def export_excel(self):
        if not self.resultados:
            messagebox.showwarning("Advertencia", "No hay resultados para exportar.")
            return
        df = pd.DataFrame(self.resultados)

        if "Contraseña" in df.columns:
            df = df.drop(columns=["Contraseña"])

        df["Fecha Ejecución"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Resultados")
                worksheet = writer.sheets["Resultados"]

                worksheet.sheet_format.defaultRowHeight = 15

                if df.columns.size:
                    for idx, column in enumerate(df.columns, start=1):
                        column_letter = worksheet.cell(row=1, column=idx).column_letter
                        worksheet.column_dimensions[column_letter].width = 48
                        for cell in worksheet[column_letter]:
                            current_alignment = cell.alignment or Alignment()
                            cell.alignment = Alignment(
                                horizontal=current_alignment.horizontal,
                                vertical=current_alignment.vertical,
                                wrap_text=True,
                            )

            messagebox.showinfo("Éxito", f"Datos exportados a {save_path}")

    def separar_pdfs(self):
        try:
            if not PYPDF2_AVAILABLE:
                messagebox.showerror("Error", "La funcionalidad para separar PDFs requiere PyPDF2.")
                return

            folder = self.folder_path.get()
            if not folder:
                messagebox.showwarning("Advertencia", "Seleccione la carpeta que contiene el PDF.")
                return

            pdf_files = [f for f in os.listdir(folder) if f.lower().endswith(".pdf")]
            if not pdf_files:
                messagebox.showwarning("Advertencia", "No se encontraron archivos PDF en la carpeta seleccionada.")
                return

            pdf_path = os.path.join(folder, pdf_files[0])

            paginas_por_archivo = simpledialog.askinteger(
                "Separar PDF", "¿Cada cuántas páginas desea dividir el PDF?", minvalue=1
            )
            if not paginas_por_archivo:
                return

            output_folder = os.path.join(folder, "PDFs_divididos")
            os.makedirs(output_folder, exist_ok=True)

            lector = PdfReader(pdf_path)
            if getattr(lector, "is_encrypted", False):
                password = self.get_password_for_file(os.path.basename(pdf_path))
                if password:
                    lector.decrypt(password)
                else:
                    messagebox.showerror("Error", "El PDF está protegido y no se encontró una contraseña registrada.")
                    return

            total_paginas = len(lector.pages)
            contador = 1

            for i in range(0, total_paginas, paginas_por_archivo):
                escritor = PdfWriter()
                for j in range(i, min(i + paginas_por_archivo, total_paginas)):
                    escritor.add_page(lector.pages[j])

                nombre_salida = f"{os.path.splitext(os.path.basename(pdf_path))[0]} {contador}.pdf"
                ruta_salida = os.path.join(output_folder, nombre_salida)
                with open(ruta_salida, "wb") as salida:
                    escritor.write(salida)
                contador += 1

            messagebox.showinfo(
                "Completado",
                f"✅ PDF dividido correctamente en {contador - 1} partes.\nGuardado en:\n{output_folder}",
            )

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo dividir el PDF:\n{e}")

    def renombrar_pdfs(self):
        try:
            excel_path = filedialog.askopenfilename(
                title="Seleccionar Excel procesado",
                filetypes=[("Archivos Excel", "*.xlsx *.xls")]
            )
            if not excel_path:
                return

            df = pd.read_excel(excel_path, dtype=str)
            df.columns = df.columns.str.strip()
            columnas = list(df.columns)

            folder = filedialog.askdirectory(title="Seleccionar carpeta con los PDFs divididos")
            if not folder:
                return

            select_win = tk.Toplevel(self.root)
            select_win.title("Seleccionar columnas para renombrar")

            tk.Label(select_win, text="Seleccione las columnas que desea usar para renombrar:").pack(pady=10)

            vars_col = {}
            for col in columnas:
                var = tk.BooleanVar()
                chk = tk.Checkbutton(select_win, text=col, variable=var)
                chk.pack(anchor="w")
                vars_col[col] = var

            def ejecutar_renombrado():
                seleccionadas = [c for c, v in vars_col.items() if v.get()]
                if not seleccionadas:
                    messagebox.showwarning("Advertencia", "Debe seleccionar al menos una columna para renombrar.")
                    return
                self._aplicar_renombrado(folder, df, seleccionadas)
                select_win.destroy()

            tk.Button(select_win, text="Renombrar PDFs", command=ejecutar_renombrado).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo iniciar el renombrado:\n{e}")

    def _aplicar_renombrado(self, folder, df, columnas):
        try:
            archivos = [f for f in os.listdir(folder) if f.lower().endswith(".pdf")]

            df["Archivo_normalizado"] = (
                df["Archivo"].astype(str).str.strip().str.replace(".pdf", "", case=False)
            ).str.lower()

            for file in archivos:
                nombre_base = os.path.splitext(file)[0].strip().lower()

                fila = df[df["Archivo_normalizado"] == nombre_base]
                if fila.empty:
                    continue

                valores = []
                for col in columnas:
                    val = str(fila.iloc[0][col]).strip()
                    val = re.sub(r'[\\/*?:"<>|]', '', val)
                    if val:
                        valores.append(val)

                if not valores:
                    continue

                nuevo_nombre = " ".join(valores) + ".pdf"
                ruta_original = os.path.join(folder, file)
                ruta_nueva = os.path.join(folder, nuevo_nombre)

                os.rename(ruta_original, ruta_nueva)

            messagebox.showinfo("Completado", f"✅ Renombrado finalizado en {folder}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al renombrar PDFs:\n{e}")

    def open_config_window(self):
        ConfigWindow(self)

    def save_config(self):
        with open(CONFIG_FILE, "wb") as f:
            pickle.dump(self.aseguradoras, f)

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "rb") as f:
                    self.aseguradoras = pickle.load(f)
            except Exception:
                self.aseguradoras = {}
        if not self.aseguradoras:
            self.aseguradoras = {"VIDA": {"Allianz": {}}}
            self.save_config()


class ConfigWindow:
    def __init__(self, app):
        self.app = app
        self.window = tk.Toplevel(app.root)
        self.window.title("Configurar Regex")

        tk.Label(self.window, text="Producto:").pack()
        self.cb_producto = ttk.Combobox(self.window, values=list(app.aseguradoras.keys()), state="readonly")
        self.cb_producto.pack(pady=5)

        if app.selected_producto.get():
            self.cb_producto.set(app.selected_producto.get())
        elif app.aseguradoras:
            self.cb_producto.set(list(app.aseguradoras.keys())[0])

        self.cb_producto.bind("<<ComboboxSelected>>", lambda e: self.refresh_aseguradoras())

        frame_prod = tk.Frame(self.window)
        frame_prod.pack(pady=5)
        tk.Button(frame_prod, text="Agregar Producto", command=self.add_producto).pack(side="left", padx=2)
        tk.Button(frame_prod, text="Editar Producto", command=self.edit_producto).pack(side="left", padx=2)
        tk.Button(frame_prod, text="Eliminar Producto", command=self.del_producto).pack(side="left", padx=2)

        tk.Label(self.window, text="Aseguradora:").pack()
        self.cb_aseg = ttk.Combobox(self.window, state="readonly")
        self.cb_aseg.pack(pady=5)
        self.cb_aseg.bind("<<ComboboxSelected>>", lambda e: self.refresh_patterns())

        frame_aseg = tk.Frame(self.window)
        frame_aseg.pack(pady=5)
        tk.Button(frame_aseg, text="Agregar Aseguradora", command=self.add_aseg).pack(side="left", padx=2)
        tk.Button(frame_aseg, text="Editar Aseguradora", command=self.edit_aseg).pack(side="left", padx=2)
        tk.Button(frame_aseg, text="Eliminar Aseguradora", command=self.del_aseg).pack(side="left", padx=2)

        self.listbox = tk.Listbox(self.window, width=100)
        self.listbox.pack(padx=5, pady=5)

        frame_pat = tk.Frame(self.window)
        frame_pat.pack(pady=5)
        tk.Button(frame_pat, text="Agregar Patrón", command=self.add_pat).pack(side="left", padx=2)
        tk.Button(frame_pat, text="Editar Patrón", command=self.edit_pat).pack(side="left", padx=2)
        tk.Button(frame_pat, text="Eliminar Patrón", command=self.del_pat).pack(side="left", padx=2)

        tk.Button(self.window, text="Guardar", command=self.save).pack(pady=5)

        self.refresh_aseguradoras()

        if app.selected_aseguradora.get():
            self.cb_aseg.set(app.selected_aseguradora.get())
            self.refresh_patterns()

    def refresh_aseguradoras(self):
        producto = self.cb_producto.get()
        if producto in self.app.aseguradoras:
            self.cb_aseg["values"] = list(self.app.aseguradoras[producto].keys())
            if self.app.aseguradoras[producto]:
                if self.app.selected_aseguradora.get() in self.app.aseguradoras[producto]:
                    self.cb_aseg.set(self.app.selected_aseguradora.get())
                else:
                    self.cb_aseg.set(list(self.app.aseguradoras[producto].keys())[0])
            else:
                self.cb_aseg.set("")
            self.refresh_patterns()
        else:
            self.cb_aseg["values"] = []
            self.cb_aseg.set("")
            self.listbox.delete(0, tk.END)

    def refresh_patterns(self):
        self.listbox.delete(0, tk.END)
        producto, aseg = self.cb_producto.get(), self.cb_aseg.get()
        if producto and aseg in self.app.aseguradoras.get(producto, {}):
            for k, v in self.app.aseguradoras[producto][aseg].items():
                self.listbox.insert(tk.END, f"{k} -> {v}")

    def add_producto(self):
        name = simpledialog.askstring("Nuevo Producto", "Nombre del producto:")
        if name and name not in self.app.aseguradoras:
            self.app.aseguradoras[name] = {}
            self.cb_producto["values"] = list(self.app.aseguradoras.keys())
            self.cb_producto.set(name)
            self.refresh_aseguradoras()
            self.app.refresh_productos_cb()

    def edit_producto(self):
        prod = self.cb_producto.get()
        if not prod:
            return
        new_name = simpledialog.askstring("Editar Producto", "Nuevo nombre:", initialvalue=prod)
        if new_name and new_name != prod:
            self.app.aseguradoras[new_name] = self.app.aseguradoras.pop(prod)
            self.cb_producto["values"] = list(self.app.aseguradoras.keys())
            self.cb_producto.set(new_name)
            self.refresh_aseguradoras()
            self.app.refresh_productos_cb()

    def del_producto(self):
        prod = self.cb_producto.get()
        if prod and messagebox.askyesno("Confirmar", f"Eliminar producto {prod}?"):
            del self.app.aseguradoras[prod]
            self.cb_producto["values"] = list(self.app.aseguradoras.keys())
            if self.app.aseguradoras:
                self.cb_producto.set(list(self.app.aseguradoras.keys())[0])
            else:
                self.cb_producto.set("")
            self.refresh_aseguradoras()
            self.app.refresh_productos_cb()

    def add_aseg(self):
        prod = self.cb_producto.get()
        if not prod:
            return
        name = simpledialog.askstring("Nueva Aseguradora", "Nombre:")
        if name and name not in self.app.aseguradoras[prod]:
            self.app.aseguradoras[prod][name] = {}
            self.cb_aseg["values"] = list(self.app.aseguradoras[prod].keys())
            self.cb_aseg.set(name)
            self.refresh_patterns()
            self.app.refresh_aseguradoras_cb()

    def edit_aseg(self):
        prod, aseg = self.cb_producto.get(), self.cb_aseg.get()
        if not prod or not aseg:
            return
        new_name = simpledialog.askstring("Editar Aseguradora", "Nuevo nombre:", initialvalue=aseg)
        if new_name and new_name != aseg:
            self.app.aseguradoras[prod][new_name] = self.app.aseguradoras[prod].pop(aseg)
            self.cb_aseg["values"] = list(self.app.aseguradoras[prod].keys())
            self.cb_aseg.set(new_name)
            self.refresh_patterns()
            self.app.refresh_aseguradoras_cb()

    def del_aseg(self):
        prod, aseg = self.cb_producto.get(), self.cb_aseg.get()
        if prod and aseg and messagebox.askyesno("Confirmar", f"Eliminar aseguradora {aseg}?"):
            del self.app.aseguradoras[prod][aseg]
            self.cb_aseg["values"] = list(self.app.aseguradoras[prod].keys())
            if self.app.aseguradoras[prod]:
                self.cb_aseg.set(list(self.app.aseguradoras[prod].keys())[0])
            else:
                self.cb_aseg.set("")
            self.refresh_patterns()
            self.app.refresh_aseguradoras_cb()

    def add_pat(self):
        prod, aseg = self.cb_producto.get(), self.cb_aseg.get()
        if not prod or not aseg:
            return
        campo = simpledialog.askstring("Nuevo Patrón", "Nombre del campo:")
        regex = simpledialog.askstring("Nuevo Patrón", "Regex:")
        if campo and regex:
            self.app.aseguradoras[prod][aseg][campo] = regex
            self.refresh_patterns()

    def edit_pat(self):
        prod, aseg = self.cb_producto.get(), self.cb_aseg.get()
        sel = self.listbox.curselection()
        if not prod or not aseg or not sel:
            return
        key = list(self.app.aseguradoras[prod][aseg].keys())[sel[0]]
        new_regex = simpledialog.askstring("Editar Patrón", f"Regex para {key}:", initialvalue=self.app.aseguradoras[prod][aseg][key])
        if new_regex:
            self.app.aseguradoras[prod][aseg][key] = new_regex
            self.refresh_patterns()

    def del_pat(self):
        prod, aseg = self.cb_producto.get(), self.cb_aseg.get()
        sel = self.listbox.curselection()
        if not prod or not aseg or not sel:
            return
        key = list(self.app.aseguradoras[prod][aseg].keys())[sel[0]]
        if messagebox.askyesno("Confirmar", f"Eliminar patrón {key}?"):
            del self.app.aseguradoras[prod][aseg][key]
            self.refresh_patterns()

    def save(self):
        self.app.save_config()
        messagebox.showinfo("Guardado", "Configuración guardada correctamente.")
        self.window.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = RegexExtractorApp(root)
    root.mainloop()
